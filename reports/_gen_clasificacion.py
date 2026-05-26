"""Generate clasificacion_papers.xlsx from verified run_twowayfe.do data."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# All 24 papers verified against each run_twowayfe.do and classify_results.txt
papers = [
    # (Paper, Journal, Wave, G, T, D, Y, Design, Subtype)
    ("Algan and Cahuc (2010)",        "AER 100(5)",  "2010-2012", "cty_num",      "period_num",    "trustgss",          "gdpk_diffswd_good",   "Other",  ""),
    ("Zhang and Zhu (2011)",          "AER 101(4)",  "2010-2012", "id",           "week_num",      "after",             "logTotal",            "Other",  ""),
    ("Bagwell and Staiger (2011)",    "AER 101(4)",  "2010-2012", "HS2",          "country_num",   "Import_M",          "TariffFinal",         "SFSD",   "SSFSD"),
    ("Wang (2011)",                   "AER 101(5)",  "2010-2012", "province_num", "year",          "post_mismatch",     "logapt_sqm",          "SFSD",   "SSFSD"),
    ("Duranton and Turner (2011)",    "AER 101(6)",  "2010-2012", "msa",          "year",          "Dl_ln_IH",          "Dl_vmt_IH",           "Other",  ""),
    ("Moser and Voena (2012)",        "AER 102(1)",  "2010-2012", "class_id",     "grntyr",        "treat",             "count_usa",           "CLA",    ""),
    ("Enikolopov et al. (2011)",      "AER 101(7)",  "2010-2012", "tik_id",       "_j",            "Watch_probit_",     "Votes_SPS_",          "HAD",    ""),
    ("Forman et al. (2012)",          "AER 102(1)",  "2010-2012", "county_id",    "period",        "surv_deeppost00",   "wagediff",            "HAD",    ""),
    ("Acemoglu et al. (2011)",        "AER 101(7)",  "2010-2012", "id",           "year",          "fpresence",         "urbrate",             "Other",  ""),
    ("Hornbeck (2012)",               "AER 102(4)",  "2010-2012", "fips",         "year",          "D_high_post",       "value_landbuildings_f","HAD",   ""),
    ("Besley and Mueller (2012)",     "AER 102(2)",  "2010-2012", "region",       "time",          "L1_wtd",            "lnhouseprice",        "SFSD",   "SSFSD"),
    ("Simcoe (2012)",                 "AER 102(1)",  "2010-2012", "techarea",     "pubCohort",     "st_stbafl1yr",      "ttlDur",              "SFSD",   "SSFSD"),
    ("Dinkelman (2011)",              "AER 101(7)",  "2010-2012", "comm_id",      "year",          "D",                 "prop_emp_f",          "CLA",    ""),
    ("Gentzkow et al. (2011)",        "AER 101(7)",  "2010-2012", "cnty90",       "year",          "numdailies",        "prestout",            "SFSD",   "SSFSD"),
    ("Baum-Snow and Lutz (2011)",     "AER 101(7)",  "2010-2012", "leaid",        "year",          "imp_post",          "lnwpu",               "SAD",    ""),
    ("Faye and Niehaus (2012)",       "AER 102(7)",  "2010-2012", "pair_id",      "year",          "i_elecex",          "oda",                 "SFSD",   "SSFSD"),
    ("Antecol et al. (2018)",         "AER 108(9)",  "2015-2019", "pol_u",        "pol_job_start", "gncs",              "tenure_policy_school", "SFSD",  "SSFSD"),
    ("Berman et al. (2017)",          "AER 107(6)",  "2015-2019", "cell",         "it",            "main_lprice_mines", "acled",               "SFSD",   "SSFSD"),
    ("Burgess et al. (2015)",         "AER 105(6)",  "2015-2019", "distnum",      "year",          "president",         "exp_dens_share",      "Other",  ""),
    ("Donaldson (2018)",              "AER 108(4-5)","2015-2019", "distid",       "year",          "RAIL",              "ln_realincome",       "SAD",    ""),
    ("Favara and Imbs (2015)",        "AER 105(3)",  "2015-2019", "county",       "year",          "Linter_bra",        "Dl_nloans_b",         "SFSD",   "SSFSD"),
    ("Fetzer (2019)",                 "AER 109(11)", "2015-2019", "id",           "ryr",           "temp",              "pct_votes_UKIP",      "SFSD",   "SSFSD"),
    ("Kaur (2019)",                   "AER 109(10)", "2015-2019", "dist",         "year",          "amons80",           "lwage",               "SFSD",   "SSFSD"),
    ("Suarez Serrato and Zidar (2016)","AER 106(9)", "2015-2019", "fe_group",     "year",          "d_keeprate",        "E",                   "Other",  ""),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Clasificacion"

headers = ["Paper", "Journal", "Wave", "G", "T", "D", "Y", "Design", "Subtype"]

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Design color coding
design_colors = {
    "CLA":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "SAD":   PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "SFSD":  PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "HAD":   PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Other": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

# Write headers
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Write data
for row_idx, paper in enumerate(papers, 2):
    for col_idx, val in enumerate(paper, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left")
        # Color the Design column
        if col_idx == 8:
            fill = design_colors.get(val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        if col_idx == 9:
            cell.alignment = Alignment(horizontal="center")

# Column widths
col_widths = [35, 14, 12, 16, 16, 20, 24, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:I{len(papers) + 1}"

# Summary sheet
ws2 = wb.create_sheet("Resumen")
ws2.cell(row=1, column=1, value="Design").font = Font(bold=True)
ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)

counts = {}
for p in papers:
    d = p[7]
    counts[d] = counts.get(d, 0) + 1

for i, (design, count) in enumerate(sorted(counts.items()), 2):
    ws2.cell(row=i, column=1, value=design)
    ws2.cell(row=i, column=2, value=count)
    fill = design_colors.get(design)
    if fill:
        ws2.cell(row=i, column=1).fill = fill

ws2.cell(row=len(counts) + 2, column=1, value="Total").font = Font(bold=True)
ws2.cell(row=len(counts) + 2, column=2, value=sum(counts.values())).font = Font(bold=True)

ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 10

outpath = r"C:\Users\Usuario\Documents\GitHub\twfe_survey\reports\clasificacion_papers.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
print(f"Total papers: {len(papers)}")
print("Counts:", counts)
