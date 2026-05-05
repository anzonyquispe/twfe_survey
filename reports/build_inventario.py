"""Build inventario_replicabilidad.xlsx — FULL PAPER audit (not just TWFE table)."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Color fills ──
VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AZUL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GRIS = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
bold = Font(bold=True, size=10)
normal = Font(size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ══════════════════════════════════════════════════════════════════
# PAPER DATA — Based on full .do file audit (May 2026)
# Format: (ola, ref, journal, data_local, ruta, url,
#          tab_total, tab_replicable, tab_no_replicable,
#          fig_total, fig_replicable,
#          estado, replicable, datasets_faltantes, razon, prioridad)
#
# estado: "Replicado full", "Replicado parcial", "Pendiente", "Sin data", "No TWFE"
# replicable: "SÍ", "NO", "PARCIAL"
# prioridad: 1=fácil (data lista, path OK), 2=necesita fix path, 3=parcial, 4=imposible
# ══════════════════════════════════════════════════════════════════

papers = []

# ═══════════════════════════════════════════════════════════
# WAVE 1: AER 2010-2012
# ═══════════════════════════════════════════════════════════

# --- Papers with full/ subdirectory (comprehensive replication) ---
papers.append(("2010-2012", "Aaronson et al. (2012)", "AER", "Sí",
    "replications/2010-2012/Aaronson et al. (2012)/full/", "",
    3, 3, 0, 0, 0,
    "Replicado full", "SÍ", "", "CPS tables replicados (Table 1, A1, A3). CEX/SIPP no intentado.", 1))

papers.append(("2010-2012", "Anderson and Sallee (2011)", "AER", "Parcial",
    "replications/2010-2012/Anderson and Sallee (2011)/full/", "",
    8, 3, 5, 6, 5,
    "Replicado parcial", "PARCIAL", "transactions1.dta (proprietary)",
    "Tables 3-7 y Figure 6 requieren datos propietarios", 3))

papers.append(("2010-2012", "Bloom et al. (2012)", "AER", "Parcial",
    "replications/2010-2012/Bloom et al. (2012)/full/", "",
    4, 4, 0, 2, 2,
    "Replicado parcial", "PARCIAL", "adib_data_10.dta (ONS restricted)",
    "Solo European results. Tables 1-5, A2, A9, C1 requieren UK data", 3))

# --- Papers with only run_twowayfe.do (dCDH analysis only) ---
papers.append(("2010-2012", "Acemoglu et al. (2011)", "AER", "Sí",
    "data/2010-2012/Acemoglu et al. (2011)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong (papers_economic→twfe_survey/data). Data presente.", 2))

papers.append(("2010-2012", "Algan and Cahuc (2010)", "AER", "Sí",
    "data/2010-2012/Algan and Cahuc (2010)/", "",
    3, 3, 0, 1, 1,
    "Pendiente", "SÍ", "", "Path wrong. Data presente (AER_MACRO.dta).", 2))

papers.append(("2010-2012", "Bagwell and Staiger (2011)", "AER", "Sí",
    "data/2010-2012/Bagwell and Staiger (2011)/", "",
    1, 1, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. Data = MainData.txt (import delimited).", 2))

papers.append(("2010-2012", "Baum-Snow and Lutz (2011)", "AER", "Sí",
    "data/2010-2012/Baum-Snow and Lutz (2011)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path correcto. dis70panx.dta presente.", 1))

papers.append(("2010-2012", "Besley and Mueller (2012)", "AER", "Sí",
    "data/2010-2012/Besley and Mueller (2012)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path correcto. maindata.dta + tourismandkillings.dta.", 1))

papers.append(("2010-2012", "Dahl and Lochner (2012)", "AER", "Sí",
    "data/2010-2012/Dahl and Lochner (2012)/", "",
    0, 0, 0, 0, 0,
    "Pendiente", "SÍ", "", "Data presente (main.dta) pero NO hay .do de replicación (solo check_vars.do).", 2))

papers.append(("2010-2012", "Dinkelman (2011)", "AER", "Sí",
    "data/2010-2012/Dinkelman (2011)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path correcto. matched_censusdata.dta.", 1))

papers.append(("2010-2012", "Duranton and Turner (2011)", "AER", "Sí",
    "data/2010-2012/Duranton and Turner (2011)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. Data presente.", 2))

papers.append(("2010-2012", "Enikolopov et al. (2011)", "AER", "Sí",
    "data/2010-2012/Enikolopov et al. (2011)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. NTV_Aggregate_Data.dta presente.", 2))

papers.append(("2010-2012", "Fang and Gavazza (2011)", "AER", "No",
    "", "",
    0, 0, 0, 0, 0,
    "No TWFE", "NO", "", "Paper usa IV/GMM, no TWFE estándar.", 4))

papers.append(("2010-2012", "Faye and Niehaus (2012)", "AER", "Sí",
    "data/2010-2012/Faye and Niehaus (2012)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path correcto. Datos completos.", 1))

papers.append(("2010-2012", "Forman et al. (2012)", "AER", "Sí",
    "data/2010-2012/Forman et al. (2012)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. countygrowth.dta presente.", 2))

papers.append(("2010-2012", "Gentzkow et al. (2011)", "AER", "Sí",
    "data/2010-2012/Gentzkow et al. (2011)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path correcto. voting_cnty_clean.dta + custom ados.", 1))

papers.append(("2010-2012", "Hornbeck (2012)", "AER", "Sí",
    "data/2010-2012/Hornbeck (2012)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. DustBowl_All_base1910.dta presente.", 2))

papers.append(("2010-2012", "Moser and Voena (2012)", "AER", "Sí",
    "data/2010-2012/Moser and Voena (2012)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. chem_patents_maindataset.dta.", 2))

papers.append(("2010-2012", "Simcoe (2012)", "AER", "Sí",
    "data/2010-2012/Simcoe (2012)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path correcto. idLevel.dta presente.", 1))

papers.append(("2010-2012", "Wang (2011)", "AER", "Sí",
    "data/2010-2012/Wang (2011)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. data_aersubmit.dta presente.", 2))

papers.append(("2010-2012", "Zhang and Zhu (2011)", "AER", "Sí",
    "data/2010-2012/Zhang and Zhu (2011)/", "",
    3, 3, 0, 0, 0,
    "Pendiente", "SÍ", "", "Path wrong. 16 .dta files + genweek.do presentes.", 2))

# --- Papers sin data (confirmed) ---
papers.append(("2010-2012", "Hotz and Xiao (2011)", "AER", "No",
    "", "",
    0, 0, 0, 0, 0,
    "Sin data", "NO", "fully_merged.dta (Census RDC confidential)",
    "Datos solo accesibles via Research Data Center", 4))

# --- Papers sin data (from original inventario, no local data) ---
sin_data_2010 = [
    ("2010-2012", "Aguiar and Hurst (2013)", "AER"),
    ("2010-2012", "Autor and Dorn (2013)", "AER"),
    ("2010-2012", "Azoulay et al. (2010)", "AER"),
    ("2010-2012", "Becker et al. (2010)", "AER"),
    ("2010-2012", "Bustos (2011)", "AER"),
    ("2010-2012", "Campante and Do (2014)", "AER"),
    ("2010-2012", "Chetty et al. (2013)", "AER"),
    ("2010-2012", "Draca et al. (2011)", "AER"),
    ("2010-2012", "Hoynes et al. (2016)", "AER"),
    ("2010-2012", "Lindo (2011)", "AER"),
    ("2010-2012", "Nunn and Qian (2011)", "AER"),
]
for ola, ref, journal in sin_data_2010:
    papers.append((ola, ref, journal, "No", "", "",
        0, 0, 0, 0, 0,
        "Sin data", "NO", "Paquete no descargado o requiere datos restringidos",
        "No disponible localmente", 4))

# ═══════════════════════════════════════════════════════════
# WAVE 2: AER 2015-2019
# ═══════════════════════════════════════════════════════════

# --- Handley & Limao: ONLY TRUE FULL REPLICATION ---
papers.append(("2015-2019", "Handley and Limao (2017)", "AER", "Sí",
    "replications/2015-2019/Handley and Limao (2017)/full/", "",
    17, 17, 0, 10, 10,
    "Replicado full", "SÍ", "",
    "FULL paper. 12 .dta, 17 tables, 10 figures. Compilado PDF.", 1))

# --- Dell: full/ directory, comprehensive ---
papers.append(("2015-2019", "Dell (2015)", "AER", "Sí",
    "replications/2015-2019/Dell (2015)/full/", "",
    3, 3, 0, 0, 0,
    "Replicado full", "SÍ", "",
    "Tables 1, 2B, 3B replicadas. Drug-specific cols redactados.", 1))

# --- Munshi: full/ directory, mostly complete ---
papers.append(("2015-2019", "Munshi and Rosenzweig (2016)", "AER", "Parcial",
    "replications/2015-2019/Munshi and Rosenzweig (2016)/full/", "",
    2, 2, 0, 0, 0,
    "Replicado parcial", "PARCIAL", "table8b.dta (para últimas 2 filas)",
    "Table 6 completa. Table 8a parcial (cgmwildboot incompatible Stata 18).", 3))

# --- Papers with run_twowayfe.do only ---
papers.append(("2015-2019", "Antecol et al. (2018)", "AER", "Sí",
    "data/2015-2019/Antecol et al. (2018)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "aer_primarysample.dta. Path correcto.", 1))

papers.append(("2015-2019", "Berman et al. (2017)", "AER", "Sí",
    "data/2015-2019/Berman et al. (2017)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "BCRT_baseline.dta. Path correcto.", 1))

papers.append(("2015-2019", "Burgess et al. (2015)", "AER", "Sí",
    "data/2015-2019/Burgess et al. (2015)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "kenya_roads_exp.dta. Path correcto.", 1))

papers.append(("2015-2019", "Donaldson (2018)", "AER", "Sí",
    "data/2015-2019/Donaldson (2018)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "income.dta + RAIL-dummies.dta. Path correcto.", 1))

papers.append(("2015-2019", "Favara and Imbs (2015)", "AER", "Sí",
    "data/2015-2019/Favara and Imbs (2015)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "hmda.dta + hp_dereg_controls.dta + call.dta.", 1))

papers.append(("2015-2019", "Fetzer (2019)", "AER", "Sí",
    "data/2015-2019/Fetzer (2019)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "DISTRICT.dta. Path correcto.", 1))

papers.append(("2015-2019", "Kaur (2019)", "AER", "Sí",
    "data/2015-2019/Kaur (2019)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "data_wb_replication.dta. Path correcto.", 1))

papers.append(("2015-2019", "Suárez Serrato and Zidar (2016)", "AER", "Sí",
    "data/2015-2019/Suárez Serrato and Zidar (2016)/", "",
    2, 2, 0, 0, 0,
    "Pendiente", "SÍ", "", "Table4.dta + muchos más. Path correcto.", 1))

# --- Sin data 2015-2019 ---
papers.append(("2015-2019", "Diamond et al. (2019)", "AER", "No",
    "", "",
    0, 0, 0, 0, 0,
    "Sin data", "NO", "infutor_panel_treat_1994_cleaned.dta (Infutor proprietary)",
    "Datos propietarios Infutor. Census/housing parciales no sirven.", 4))

papers.append(("2015-2019", "Pierce and Schott (2016)", "AER", "No",
    "", "",
    0, 0, 0, 0, 0,
    "Sin data", "NO", "Census Longitudinal Business Database (restricted)",
    "Requiere Census Research Data Center", 4))

papers.append(("2015-2019", "Allcott et al. (2016)", "AER", "No",
    "", "",
    0, 0, 0, 0, 0,
    "Sin data", "NO", "ASI plant microdata (restricted India)",
    "Annual Survey of Industries microdata confidencial", 4))

# Sin data adicionales 2015-2019
sin_data_2015 = [
    ("2015-2019", "Aghion et al. (2019)", "AER"),
    ("2015-2019", "Autor et al. (2017)", "AER"),
    ("2015-2019", "Caliendo et al. (2019)", "AER"),
    ("2015-2019", "Dix-Carneiro and Kovak (2017)", "AER"),
    ("2015-2019", "Goldsmith-Pinkham et al. (2020)", "AER"),
]
for ola, ref, journal in sin_data_2015:
    papers.append((ola, ref, journal, "No", "", "",
        0, 0, 0, 0, 0,
        "Sin data", "NO", "Paquete no descargado o requiere datos restringidos",
        "No disponible localmente", 4))

# ═══════════════════════════════════════════════════════════
# WAVE 3: Political Science (11 papers)
# ═══════════════════════════════════════════════════════════

polisci_base = "replication_did_info/dataverse_files/replication/replication/STATA/"
polisci_papers = [
    ("Bischof and Wagner (2019)", "AJPS", "01_bischof.do", "data_bischof.dta", ""),
    ("Clayton and Zetterberg (2018)", "JOP", "02_clayton.do", "data_clayton.dta", "Controls: demlag, dem_cumlag, etc."),
    ("Christensen and Garfias (2021)", "JOP", "03_christensen.do", "data_christensen.dta", "Submuestra chris_sub."),
    ("Hankinson and Magazinnik (2023)", "JOP", "04_hankinson.do", "data_hankinson.dta", "Controls: mht, population, etc."),
    ("Eckhouse (2022)", "AJPS", "05_eckhouse.do", "data_eckhouse.dta", "drop.period logic. CS Never NA."),
    ("Paglayan (2022)", "APSR", "06_paglayan.do", "data_paglayan.dta", ""),
    ("Kuipers and Sahn (2023)", "APSR", "07_kuipers.do", "data_kuipers.dta", "Control: pop_prop."),
    ("Esberg and Siegel (2023)", "APSR", "08_esberg.do", "data_esberg.dta", "Monthly data. SA/CS crash (too many periods)."),
    ("Hirano et al. (2022)", "JOP", "09_hirano.do", "data_hirano.dta", "Even years only. CS Never NA."),
    ("Trounstine (2020)", "APSR", "10a_trounstine.do", "data_trounstine.dta", "drop.period logic."),
    ("Hainmueller and Hangartner (2019)", "AJPS", "11a_hainmueller.do", "data_hainmueller.dta", ""),
]

for ref, journal, do_file, dta_file, notes in polisci_papers:
    # Each produces 8 estimators saved to output/*.dta (no formal LaTeX tables)
    papers.append(("PolSci", ref, journal, "Sí",
        polisci_base + dta_file, "",
        1, 1, 0, 0, 0,
        "Pendiente", "SÍ", "",
        f"Self-contained ({dta_file}). 8 estimators. {notes}", 1))

# ═══════════════════════════════════════════════════════════
# WAVE 4: Candidates (not yet downloaded)
# ═══════════════════════════════════════════════════════════

candidates = [
    ("Cengiz et al. (2019)", "QJE", "https://doi.org/10.7910/DVN/BKIYOL"),
    ("Dube, Lester and Reich (2010)", "REStat", "https://doi.org/10.7910/DVN/AKHFAZ"),
    ("Autor, Dorn and Hanson (2013)", "AER", "https://www.openicpsr.org/openicpsr/project/116439"),
    ("Baker, Larcker and Wang (2022)", "JFE", "https://github.com/andrewchbaker/DiDDesign"),
    ("Harari and La Ferrara (2018)", "REStat", "https://doi.org/10.7910/DVN/NYJCBR"),
]
for ref, journal, url in candidates:
    papers.append(("Candidato", ref, journal, "No",
        "", url,
        0, 0, 0, 0, 0,
        "No descargado", "SÍ", "",
        "Paquete verificado online. Pendiente descarga.", 2))


# ══════════════════════════════════════════════════════════════════
# SHEET 1: Inventario Completo
# ══════════════════════════════════════════════════════════════════

ws = wb.active
ws.title = "Inventario Completo"

headers = ["#", "Ola", "Referencia", "Journal", "Data local", "Ruta",
           "Data online URL", "Tab. totales", "Tab. replicables",
           "Tab. NO replicables", "Fig. totales", "Fig. replicables",
           "Estado", "Replicable", "Datasets faltantes", "Razón/Notas", "Prioridad"]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = HEADER
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Write data
for i, p in enumerate(papers, 1):
    ola, ref, journal, data_local, ruta, url, tab_tot, tab_rep, tab_no, fig_tot, fig_rep, estado, replicable, faltantes, razon, prioridad = p
    row = i + 1
    vals = [i, ola, ref, journal, data_local, ruta, url,
            tab_tot, tab_rep, tab_no, fig_tot, fig_rep,
            estado, replicable, faltantes, razon, prioridad]

    # Determine fill color
    if estado == "Replicado full" and ref == "Handley and Limao (2017)":
        fill = AZUL  # Only true FULL replication
    elif estado == "Replicado full":
        fill = AZUL  # Dell, Aaronson also fully replicated their scope
    elif estado == "Replicado parcial":
        fill = AMARILLO
    elif estado in ("Sin data", "No TWFE"):
        fill = ROJO
    elif estado == "No descargado":
        fill = GRIS
    elif estado == "Pendiente" and replicable == "SÍ":
        fill = VERDE
    else:
        fill = GRIS

    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = normal
        cell.border = thin_border
        if col in (1, 4, 5, 8, 9, 10, 11, 12, 14, 17):
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(wrap_text=True)

# Column widths
col_widths = [4, 10, 35, 7, 8, 50, 40, 8, 8, 8, 8, 8, 15, 9, 40, 55, 8]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
# SHEET 2: Cola de trabajo
# ══════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Cola de trabajo")

# Filter: only replicable papers that are "Pendiente"
cola = [(i+1, p) for i, p in enumerate(papers)
        if p[11] == "Pendiente" and p[12] == "SÍ"]

# Sort by priority then by wave
cola.sort(key=lambda x: (x[1][15], x[1][0], x[1][1]))

headers2 = ["Orden", "Ola", "Referencia", "Journal", "Tab. totales",
            "Tab. replicables", "Fig. totales", "Prioridad", "Notas"]

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = HEADER
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

for row_idx, (orig_num, p) in enumerate(cola, 2):
    ola, ref, journal, _, _, _, tab_tot, tab_rep, _, fig_tot, fig_rep, _, _, _, razon, prioridad = p
    vals = [row_idx - 1, ola, ref, journal, tab_tot, tab_rep, fig_tot, prioridad, razon]
    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.font = normal
        cell.border = thin_border
        if prioridad == 1:
            cell.fill = VERDE
        elif prioridad == 2:
            cell.fill = AMARILLO

col_widths2 = [6, 10, 35, 7, 10, 10, 10, 9, 60]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
# SHEET 3: Resumen
# ══════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("Resumen")

# Counts
n_azul = sum(1 for p in papers if p[11].startswith("Replicado full"))
n_amarillo = sum(1 for p in papers if p[11] == "Replicado parcial")
n_verde = sum(1 for p in papers if p[11] == "Pendiente" and p[12] == "SÍ")
n_rojo = sum(1 for p in papers if p[11] in ("Sin data", "No TWFE"))
n_gris = sum(1 for p in papers if p[11] == "No descargado")
n_total = len(papers)
n_cola = len(cola)

# Total tables/figures
total_tabs = sum(p[6] for p in papers)   # tab_total at index 6
total_tabs_rep = sum(p[7] for p in papers)  # tab_replicable at index 7
total_figs = sum(p[9] for p in papers)   # fig_total at index 9
total_figs_rep = sum(p[10] for p in papers)  # fig_replicable at index 10

ws3.cell(row=1, column=1, value="RESUMEN DE REPLICABILIDAD").font = Font(bold=True, size=14)
ws3.cell(row=2, column=1, value=f"Total papers: {n_total}").font = bold

summary_data = [
    ("AZUL (replicado full)", n_azul, AZUL),
    ("AMARILLO (replicado parcial)", n_amarillo, AMARILLO),
    ("VERDE (pendiente con data)", n_verde, VERDE),
    ("GRIS (no descargado)", n_gris, GRIS),
    ("ROJO (sin data / imposible)", n_rojo, ROJO),
]

for i, (label, count, fill) in enumerate(summary_data, 4):
    ws3.cell(row=i, column=1, value=label).font = bold
    ws3.cell(row=i, column=1).fill = fill
    ws3.cell(row=i, column=2, value=count).font = bold
    ws3.cell(row=i, column=2).fill = fill

ws3.cell(row=10, column=1, value="EN COLA DE TRABAJO:").font = bold
ws3.cell(row=10, column=2, value=n_cola).font = bold

ws3.cell(row=12, column=1, value="TABLAS/FIGURAS").font = Font(bold=True, size=12)
ws3.cell(row=13, column=1, value="Total tablas en .do files:").font = normal
ws3.cell(row=13, column=2, value=total_tabs).font = bold
ws3.cell(row=14, column=1, value="Tablas replicables con data disponible:").font = normal
ws3.cell(row=14, column=2, value=total_tabs_rep).font = bold

ws3.cell(row=16, column=1, value="PRIORIDAD 1 (path correcto, data lista):").font = bold
ws3.cell(row=16, column=2, value=sum(1 for p in papers if p[15] == 1 and p[11] == "Pendiente")).font = bold
ws3.cell(row=17, column=1, value="PRIORIDAD 2 (necesita fix path o .do):").font = bold
ws3.cell(row=17, column=2, value=sum(1 for p in papers if p[15] == 2 and p[11] == "Pendiente")).font = bold

ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 10


# ══════════════════════════════════════════════════════════════════
# SHEET 4: Verificación .dta
# ══════════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("Verificación .dta")

# Verification data: (paper, dta_required, dta_present, dta_missing_list, pct_complete)
# All 36 green papers verified on 2026-05-05 — ALL are 100%

verificacion = [
    # --- AER 2010-2012 (17 papers) ---
    ("Acemoglu et al. (2011)", 1, 1, [], 100,
     "20100816_replication_dataset.dta"),
    ("Algan and Cahuc (2010)", 1, 1, [], 100,
     "AER_MACRO.dta"),
    ("Bagwell and Staiger (2011)", 1, 1, [], 100,
     "MainData.txt (import delimited)"),
    ("Baum-Snow and Lutz (2011)", 1, 1, [], 100,
     "dis70panx.dta"),
    ("Besley and Mueller (2012)", 2, 2, [], 100,
     "maindata.dta, tourismandkillings.dta"),
    ("Dahl and Lochner (2012)", 1, 1, [], 100,
     "main.dta"),
    ("Dinkelman (2011)", 1, 1, [], 100,
     "matched_censusdata.dta"),
    ("Duranton and Turner (2011)", 1, 1, [], 100,
     "Duranton_Turner_AER_2010.dta"),
    ("Enikolopov et al. (2011)", 1, 1, [], 100,
     "NTV_Aggregate_Data.dta"),
    ("Faye and Niehaus (2012)", 1, 1, [], 100,
     "111102_oda_final_data_big5_commit_080107_unvotes_term.dta"),
    ("Forman et al. (2012)", 1, 1, [], 100,
     "countygrowth.dta"),
    ("Gentzkow et al. (2011)", 1, 1, [], 100,
     "voting_cnty_clean.dta + custom .ado files"),
    ("Hornbeck (2012)", 1, 1, [], 100,
     "DustBowl_All_base1910.dta"),
    ("Moser and Voena (2012)", 1, 1, [], 100,
     "chem_patents_maindataset.dta"),
    ("Simcoe (2012)", 1, 1, [], 100,
     "idLevel.dta"),
    ("Wang (2011)", 1, 1, [], 100,
     "data_aersubmit.dta"),
    ("Zhang and Zhu (2011)", 6, 6, [], 100,
     "daily_contribution.dta, userlangstat_percentage.dta, genweek.do, "
     "userpagestat2005.dta, usertalkstat2005.dta, percent_blocked.dta"),
    # --- AER 2015-2019 (8 papers) ---
    ("Antecol et al. (2018)", 1, 1, [], 100,
     "aer_primarysample.dta"),
    ("Berman et al. (2017)", 1, 1, [], 100,
     "BCRT_baseline.dta"),
    ("Burgess et al. (2015)", 1, 1, [], 100,
     "kenya_roads_exp.dta"),
    ("Donaldson (2018)", 2, 2, [], 100,
     "income.dta, RAIL-dummies.dta"),
    ("Favara and Imbs (2015)", 3, 3, [], 100,
     "hmda.dta, hp_dereg_controls.dta, call.dta"),
    ("Fetzer (2019)", 1, 1, [], 100,
     "DISTRICT.dta"),
    ("Kaur (2019)", 1, 1, [], 100,
     "data_wb_replication.dta"),
    ("Suárez Serrato and Zidar (2016)", 1, 1, [], 100,
     "Table4.dta (+ others referenced)"),
    # --- PolSci (11 papers) ---
    ("Bischof and Wagner (2019)", 1, 1, [], 100,
     "data_bischof.dta"),
    ("Clayton and Zetterberg (2018)", 1, 1, [], 100,
     "data_clayton.dta"),
    ("Christensen and Garfias (2021)", 1, 1, [], 100,
     "data_christensen.dta"),
    ("Hankinson and Magazinnik (2023)", 1, 1, [], 100,
     "data_hankinson.dta"),
    ("Eckhouse (2022)", 1, 1, [], 100,
     "data_eckhouse.dta"),
    ("Paglayan (2022)", 1, 1, [], 100,
     "data_paglayan.dta"),
    ("Kuipers and Sahn (2023)", 1, 1, [], 100,
     "data_kuipers.dta"),
    ("Esberg and Siegel (2023)", 1, 1, [], 100,
     "data_esberg.dta"),
    ("Hirano et al. (2022)", 1, 1, [], 100,
     "data_hirano.dta"),
    ("Trounstine (2020)", 1, 1, [], 100,
     "data_trounstine.dta"),
    ("Hainmueller and Hangartner (2019)", 1, 1, [], 100,
     "data_hainmueller.dta"),
]

headers4 = ["#", "Paper", "Total .dta requeridos", ".dta presentes",
            ".dta faltantes", "% Completitud", "Archivos verificados"]

for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.fill = HEADER
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

for i, (paper, req, pres, missing, pct, files) in enumerate(verificacion, 1):
    row = i + 1
    missing_str = ", ".join(missing) if missing else ""
    vals = [i, paper, req, pres, missing_str, f"{pct}%", files]
    fill = VERDE if pct == 100 else AMARILLO
    for col, val in enumerate(vals, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = normal
        cell.border = thin_border
        if col in (1, 3, 4, 6):
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(wrap_text=True)

# Summary row
sum_row = len(verificacion) + 2
ws4.cell(row=sum_row, column=1, value="").font = bold
ws4.cell(row=sum_row, column=2, value="TOTAL").font = bold
ws4.cell(row=sum_row, column=3, value=sum(v[1] for v in verificacion)).font = bold
ws4.cell(row=sum_row, column=4, value=sum(v[2] for v in verificacion)).font = bold
ws4.cell(row=sum_row, column=5, value="").font = bold
ws4.cell(row=sum_row, column=6, value="100%").font = bold
ws4.cell(row=sum_row, column=6).fill = VERDE
for col in range(1, 8):
    ws4.cell(row=sum_row, column=col).border = thin_border

col_widths4 = [4, 38, 18, 14, 20, 14, 65]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════

out_path = r"C:\Users\Usuario\Documents\GitHub\twfe_survey\reports\inventario_replicabilidad.xlsx"
wb.save(out_path)
print(f"Saved to {out_path}")
print(f"\nRESUMEN:")
print(f"  AZUL (replicado full):       {n_azul}")
print(f"  AMARILLO (replicado parcial): {n_amarillo}")
print(f"  VERDE (pendiente con data):   {n_verde}")
print(f"  GRIS (no descargado):         {n_gris}")
print(f"  ROJO (sin data/imposible):    {n_rojo}")
print(f"  TOTAL:                        {n_total}")
print(f"  EN COLA:                      {n_cola}")
print(f"\n  Tablas totales:              {total_tabs}")
print(f"  Tablas replicables:          {total_tabs_rep}")
print(f"  Figuras totales:             {total_figs}")
print(f"  Figuras replicables:         {total_figs_rep}")
print(f"\n  VERIFICACIÓN .dta:")
print(f"  Papers verificados:          {len(verificacion)}")
print(f"  Total archivos requeridos:   {sum(v[1] for v in verificacion)}")
print(f"  Total archivos presentes:    {sum(v[2] for v in verificacion)}")
print(f"  Papers 100% completos:       {sum(1 for v in verificacion if v[4]==100)}/{len(verificacion)}")
