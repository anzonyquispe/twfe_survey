"""
Generate estado_papers_final.xlsx for the TWFE survey project.
All 59 papers with detailed status, G/T/D/Y availability, and justifications.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Estado Papers"

# ============================================================================
# COLORS
# ============================================================================
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=9)
BOLD_FONT = Font(name="Calibri", size=9, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ============================================================================
# HEADERS
# ============================================================================
headers = [
    "#",
    "Paper (autores, ano)",
    "Journal",
    "Wave",
    "Tabla dCDH",
    "G (grupo)",
    "T (tiempo)",
    "D (tratamiento)",
    "Y (outcome)",
    "G disp.",
    "T disp.",
    "D disp.",
    "Y disp.",
    "Estado",
    "Nota detallada",
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

# Freeze top row
ws.freeze_panes = "A2"

# ============================================================================
# DATA — ALL 59 PAPERS
# ============================================================================

papers = [
    # ========================================================================
    # WAVE 1 (2010-2012) — 33 papers
    # ========================================================================

    # --- REPLICADO (16 papers with panel_GTD.dta in replications/) ---
    {
        "paper": "Acemoglu, Cantoni, Johnson, Robinson (2011)",
        "wave": "2010-2012",
        "tabla": "Table 3",
        "G": "id (country)", "T": "year", "D": "fpresence", "Y": "urbrate",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 75% negative weights (6 pos / 18 neg). Democracy and urbanization panel.",
    },
    {
        "paper": "Algan and Cahuc (2010)",
        "wave": "2010-2012",
        "tabla": "Figure 4 (Tables VI-VII)",
        "G": "cty_num (country)", "T": "period_num", "D": "trustgss", "Y": "gdpk_diffswd_good",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. Panel structure incompatible con TWFE weight decomposition (rc=402). Trust and growth.",
    },
    {
        "paper": "Bagwell and Staiger (2011)",
        "wave": "2010-2012",
        "tabla": "Table 3, OLS cols",
        "G": "HS2 (product)", "T": "country_num", "D": "Import_M", "Y": "TariffFinal",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 29.3% negative weights (1044 pos / 432 neg). Trade agreements and tariffs.",
    },
    {
        "paper": "Baum-Snow and Lutz (2011)",
        "wave": "2010-2012",
        "tabla": "Tables 2-6",
        "G": "leaid (district)", "T": "year", "D": "imp_post", "Y": "lnwpu",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 26.9% negative weights (152 pos / 56 neg). Desegregation staggered adoption.",
    },
    {
        "paper": "Besley and Mueller (2012)",
        "wave": "2010-2012",
        "tabla": "Table 1, Cols 3, 5-7",
        "G": "region", "T": "time", "D": "L1_wtd", "Y": "lnhouseprice",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 40.5% negative weights (184 pos / 125 neg). Conflict and investment in Northern Ireland.",
    },
    {
        "paper": "Dinkelman (2011)",
        "wave": "2010-2012",
        "tabla": "Tables 4, 5, 8, 9, 10",
        "G": "comm_id (community)", "T": "year", "D": "D (binary)", "Y": "prop_emp_f",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 0% negative weights (365 pos / 0 neg). Binary treatment x 2 periods, clean design.",
    },
    {
        "paper": "Duranton and Turner (2011)",
        "wave": "2010-2012",
        "tabla": "Table 5",
        "G": "msa", "T": "year", "D": "Dl_ln_IH", "Y": "Dl_vmt_IH",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. Panel structure incompatible con weight decomposition (rc=402). Urban highways and sprawl.",
    },
    {
        "paper": "Enikolopov, Petrova, Zhuravskaya (2011)",
        "wave": "2010-2012",
        "tabla": "Table 3",
        "G": "tik_id (region)", "T": "_j (election)", "D": "Watch_probit_", "Y": "Votes_SPS_",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 52.6% negative weights (918 pos / 1020 neg). Media and political persuasion.",
    },
    {
        "paper": "Faye and Niehaus (2012)",
        "wave": "2010-2012",
        "tabla": "Table 3, Cols 4-5; Tables 4-5",
        "G": "pair_id", "T": "year", "D": "i_elecex", "Y": "oda",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 0% negative weights (1370 pos / 0 neg). Clean design, no negative weights.",
    },
    {
        "paper": "Forman, Goldfarb, Greenstein (2012)",
        "wave": "2010-2012",
        "tabla": "Tables 2, 4",
        "G": "county_id", "T": "period", "D": "surv_deeppost00", "Y": "wagediff",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. All weights zero/missing, degenerate panel (rc=109). Internet and wage inequality.",
    },
    {
        "paper": "Gentzkow, Shapiro, Sinkinson (2011)",
        "wave": "2010-2012",
        "tabla": "Tables 2, 3",
        "G": "cnty90 (county)", "T": "year", "D": "numdailies", "Y": "prestout",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 41.1% negative weights (5919 pos / 4137 neg). Newspapers and voter turnout.",
    },
    {
        "paper": "Hornbeck (2012)",
        "wave": "2010-2012",
        "tabla": "Table 2",
        "G": "fips (county)", "T": "year", "D": "D_high_post", "Y": "value_landbuildings_f",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 32.2% negative weights (2343 pos / 1111 neg). Dust Bowl long-run effects.",
    },
    {
        "paper": "Moser and Voena (2012)",
        "wave": "2010-2012",
        "tabla": "Table 2",
        "G": "class_id (patent class)", "T": "grntyr (grant year)", "D": "treat", "Y": "count_usa",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 0% negative weights (7056 pos / 0 neg). Clean design, compulsory licensing.",
    },
    {
        "paper": "Simcoe (2012)",
        "wave": "2010-2012",
        "tabla": "Table 4, Cols 1-3",
        "G": "techarea", "T": "pubCohort", "D": "st_stbafl1yr", "Y": "ttlDur",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 50.8% negative weights (32 pos / 33 neg). Standard-setting committees.",
    },
    {
        "paper": "Wang (2011)",
        "wave": "2010-2012",
        "tabla": "Table 5, Panel A",
        "G": "province_num", "T": "year", "D": "post_mismatch", "Y": "logapt_sqm",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. Panel structure incompatible con weight decomposition (rc=402). Colonial institutions.",
    },
    {
        "paper": "Zhang and Zhu (2011)",
        "wave": "2010-2012",
        "tabla": "Tables 3-4, Cols 4-6",
        "G": "id (contributor)", "T": "week_num", "D": "after", "Y": "logTotal",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 53.9% negative weights (54246 pos / 63537 neg). Wikipedia contributions.",
    },

    # --- REPLICADO NEW (Hotz & Xiao) ---
    {
        "paper": "Hotz and Xiao (2011)",
        "wave": "2010-2012",
        "tabla": "Table 7, Cols 4-5 (nonemployer analysis)",
        "G": "num_st (state)", "T": "year (1987, 1992, 1997)", "D": "scrat (staff-child ratio), educ (education index)", "Y": "estab (nonemployer establishments)",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "Pipeline completo. 153 obs (51 estados x 3 anos). twowayfeweights feTR: 60 pos / 89 neg (59.7% negative). Design: SFSD (RSFSD). Spec: areg estab scrat educ [controls] year1992 year1997, absorb(num_st) cluster(year_st). PDF compilado.",
    },

    # --- RESTRINGIDO (Wave 1) ---
    {
        "paper": "Aizer (2010)",
        "wave": "2010-2012",
        "tabla": "Table 2",
        "G": "zip/hospital", "T": "year", "D": "ratiow_hs (Bartik wage instrument)", "Y": "lfass (log assault rate)",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "D (Bartik wages) disponible en bartik03_aer.dta. Y (hospital discharges por asalto) requiere hosp2_2003.dta de OSHPD (Office of Statewide Health Planning and Development, California). OSHPD requiere solicitud institucional, acuerdo de confidencialidad y pago (~$1000-5000). Los datos de alta hospitalaria contienen informacion de pacientes protegida por HIPAA.",
    },
    {
        "paper": "Anderson and Sallee (2011)",
        "wave": "2010-2012",
        "tabla": "Table 5, Col 2",
        "G": "vehicle model", "T": "year", "D": "CAFE compliance", "Y": "price/quantity",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "cafe_compliance.dta disponible con variables de regulacion CAFE. Falta transactions1.dta con datos de ventas de vehiculos de J.D. Power and Associates. J.D. Power (ahora JD Power/S&P Global Mobility) vende datos de transacciones a nivel VIN; licencia comercial ~$50K-100K/ano. Requiere acuerdo institucional.",
    },
    {
        "paper": "Bajari, Fruehwirth, Kim, Timmins (2012)",
        "wave": "2010-2012",
        "tabla": "Table 5",
        "G": "property/tract", "T": "year", "D": "pollution changes (TSP, ozone)", "Y": "housing prices",
        "G_d": "NO", "T_d": "NO", "D_d": "SI", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Solo datos de contaminacion presentes (Census tract pollution). Faltan datos de precios de vivienda de DataQuick (ahora CoreLogic). CoreLogic proporciona datos de transacciones inmobiliarias bajo licencia comercial (~$10K-50K/ano dependiendo de cobertura). Se necesita acuerdo institucional de investigacion. G y T no identificables sin datos de transacciones.",
    },
    {
        "paper": "Bloom, Sadun, Van Reenen (2012)",
        "wave": "2010-2012",
        "tabla": "Table 2, Cols 6-8",
        "G": "firm/establishment", "T": "year", "D": "management practices", "Y": "productivity",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Solo hay archivos CSV parciales en European-Results/. Los datos principales son del UK Census of Production y Annual Respondents Database del ONS (Office for National Statistics). Acceso requiere presentar proyecto al UK Secure Research Service (SRS), antes Virtual Microdata Laboratory (VML), en Londres. Proceso tarda ~3-6 meses; investigador debe ser de institucion acreditada en UK.",
    },
    {
        "paper": "Brambilla, Lederman, Porto (2012)",
        "wave": "2010-2012",
        "tabla": "Table 5",
        "G": "firm", "T": "year", "D": "export destination quality", "Y": "wages/skill premium",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Todas las variables provienen de firmdata (Encuesta Industrial Anual, EIA, recopilada por INDEC Argentina). Datos protegidos por Ley 17.622 de Secreto Estadistico de Argentina. Solo accesibles a investigadores argentinos con autorizacion del INDEC. No hay mecanismo formal para acceso internacional. Los 77.6 MB del package son codigo .do sin datos.",
    },
    {
        "paper": "Chaney, Sraer, Thesmar (2012)",
        "wave": "2010-2012",
        "tabla": "Table 5",
        "G": "msa", "T": "year", "D": "elasticity x mortgage rate", "Y": "index_msa (house prices)",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI/PARCIAL",
        "estado": "Restringido",
        "nota": "Table 5 Cols 1-2 (first stage residencial) SI replicable: Y=index_msa de housing_data_MSA.txt, D=elasticity*mortgage, G=msa, T=year. Table 5 Cols 3-4 NO: requieren offprice (commercial real estate) de CoStar Group, licencia comercial ~$30K/ano. COMPUSTAT datos tambien requeridos para columnas de inversion corporativa (Standard & Poor's, ~$25K/ano via WRDS).",
    },
    {
        "paper": "Chandra, Gruber, McKnight (2010)",
        "wave": "2010-2012",
        "tabla": "Tables 2-3, first line",
        "G": "HRR (hospital referral region)", "T": "year", "D": "treatment intensity", "Y": "health outcomes",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Datos de Medicare claims (CMS). Requieren acceso al Research Data Assistance Center (ResDAC) de CMS. Proceso formal: Data Use Agreement (DUA), aprobacion IRB, ~$20K-100K en fees dependiendo del tamano de la extraccion. Solo .do files en el package, sin ningun .dta.",
    },
    {
        "paper": "Dafny, Duggan, Ramanarayanan (2012)",
        "wave": "2010-2012",
        "tabla": "Table 3",
        "G": "insurer-state", "T": "year", "D": "merger exposure", "Y": "premiums",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Datos de primas de seguros de salud del LEHID (Large Employer Health Insurance Dataset) de Kaiser/HRET y MEPS-IC. LEHID ya no esta disponible (encuesta descontinuada). MEPS-IC requiere acceso al Census Bureau RDC. Solo .do files en el package, sin ningun .dta.",
    },
    {
        "paper": "Duggan and Morton (2010)",
        "wave": "2010-2012",
        "tabla": "Tables 2-3",
        "G": "drug molecule", "T": "quarter", "D": "Medicare share (reformulation share)", "Y": "prices/doses",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "D (Medicare share) disponible via MEPS (merge0203AA.dta, meps02share.dta). Y (precios y volumenes de prescripcion) requiere datos de IMS Health (ahora IQVIA). IQVIA National Sales Perspectives cuesta ~$50K-200K/ano bajo licencia comercial. Algunos .dta parciales presentes (mcd2003us.dta, miscvars1.dta) pero no los datos principales de precios.",
    },
    {
        "paper": "Ellul, Pagano, Panunzi (2010)",
        "wave": "2010-2012",
        "tabla": "Table 7",
        "G": "firm/country", "T": "year", "D": "employment protection (3 treatments)", "Y": "leverage/structure",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Datos principales de WorldScope/Datastream (Thomson Reuters, ahora Refinitiv/LSEG). Licencia universitaria via Wharton Research Data Services (WRDS) ~$25K/ano. Tambien requiere datos de proteccion laboral del Banco Mundial/OECD. Solo .do files en el package, cero .dta.",
    },
    {
        "paper": "Imberman, Kugler, Sacerdote (2012)",
        "wave": "2010-2012",
        "tabla": "Tables 3-6",
        "G": "school", "T": "year", "D": "Katrina evacuee share", "Y": "test scores",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Datos de rendimiento escolar a nivel de estudiante del Houston ISD y Louisiana DOE. Datos protegidos por FERPA (Family Educational Rights and Privacy Act). Requiere acuerdo de datos con cada distrito escolar. 79 .do files en el package, cero .dta.",
    },
    {
        "paper": "Mian and Sufi (2011)",
        "wave": "2010-2012",
        "tabla": "Tables 2-3",
        "G": "zip code", "T": "year", "D": "housing net worth shock", "Y": "auto sales/employment",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Datos de Equifax (credit bureau) y Lender Processing Services (LPS, ahora Black Knight/ICE Mortgage Technology). Equifax requiere licencia comercial; LPS/Black Knight Academic Research License ~$10K-25K/ano. Tambien requiere datos de Polk (registros de vehiculos). Solo .do files, cero .dta.",
    },

    # --- FALTA DATA (Wave 1) ---
    {
        "paper": "Bustos (2011)",
        "wave": "2010-2012",
        "tabla": "Tables 3-12",
        "G": "firm", "T": "year", "D": "tariff reduction (Mercosur)", "Y": "technology adoption",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Falta data",
        "nota": "Cero archivos .dta en el package. Solo codigo .do. Los datos principales son de la Encuesta Nacional de Innovacion Tecnologica (ENIT) de INDEC Argentina, protegidos por secreto estadistico (Ley 17.622). El replication package fue subido sin datos.",
    },

    # --- OTROS (Wave 1: no TWFE, TAXSIM, structural, partial) ---
    {
        "paper": "Aaronson, Agarwal, French (2012)",
        "wave": "2010-2012",
        "tabla": "Tables 1, 2, 5",
        "G": "household/state", "T": "quarter/year", "D": "minimum wage change", "Y": "credit/spending",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Restringido",
        "nota": "Datos del Consumer Expenditure Survey (CES/CEX) del BLS disponibles publicamente, pero el codigo requiere GAUSS (software propietario) y archivos comprimidos con formato especifico. Analisis principal es structural/reduced-form, no TWFE estandar. Acceso parcial; replicacion requiere reescribir codigo de GAUSS a Stata.",
    },
    {
        "paper": "Dahl and Lochner (2012)",
        "wave": "2010-2012",
        "tabla": "Table 3",
        "G": "family", "T": "year", "D": "EITC income shock (IV)", "Y": "child test scores",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Falta data",
        "nota": "main.dta (33 MB) presente con datos NLSY. PERO la variable state=0 en todas las obs, sugiriendo datos parcialmente censurados. Ademas, el codigo requiere TAXSIM9 (software de NBER para simular impuestos), cuyo endpoint web historico ya no esta disponible. Sin TAXSIM no se pueden construir las variables instrumentales de EITC.",
    },
    {
        "paper": "Fang and Gavazza (2011)",
        "wave": "2010-2012",
        "tabla": "Tables 2, 3, 5, 6 Col 3",
        "G": "county/state", "T": "year", "D": "Medigap regulations", "Y": "insurance enrollment",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido en replications/. Modelo estructural dinamico, no TWFE estandar. Los datos estan completos (1205.6 MB). Analizado pero el diseno no es compatible con twowayfeweights.",
    },

    # ========================================================================
    # WAVE 2 (2015-2019) — 26 papers
    # ========================================================================

    # --- REPLICADO (8 papers with panel_GTD.dta) ---
    {
        "paper": "Antecol, Bedard, Stearns (2018)",
        "wave": "2015-2019",
        "tabla": "Table 2",
        "G": "pol_u (university)", "T": "pol_job_start", "D": "gncs (gender-neutral clock stopping)", "Y": "tenure_policy_school",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 5.1% negative weights (186 pos / 10 neg). Tenure clock stopping policies.",
    },
    {
        "paper": "Berman, Couttenier, Rohner, Thoenig (2017)",
        "wave": "2015-2019",
        "tabla": "Table 2, Col 2",
        "G": "cell (grid cell)", "T": "it (year)", "D": "main_lprice_mines", "Y": "acled (conflict events)",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 58.6% negative weights (597 pos / 845 neg). Mining prices and conflict in Africa.",
    },
    {
        "paper": "Burgess, Jedwab, Miguel, Morjaria, Padro i Miquel (2015)",
        "wave": "2015-2019",
        "tabla": "Table 1, Col 1",
        "G": "distnum (district)", "T": "year", "D": "president (co-ethnic president)", "Y": "exp_dens_share",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 0% negative weights (319 pos / 0 neg). Binary treatment switches on/off. Kenya roads and ethnic favoritism.",
    },
    {
        "paper": "Donaldson (2018)",
        "wave": "2015-2019",
        "tabla": "Table 4, Col 1",
        "G": "distid (district)", "T": "year", "D": "RAIL (railroad access)", "Y": "ln_realincome",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 51.0% negative weights (3169 pos / 3293 neg). Railroads and welfare in colonial India.",
    },
    {
        "paper": "Favara and Imbs (2015)",
        "wave": "2015-2019",
        "tabla": "Table 4, Col 1",
        "G": "county", "T": "year", "D": "Linter_bra (interstate branching deregulation)", "Y": "Dl_nloans_b",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 40.7% negative weights (3983 pos / 2730 neg). Credit supply and house prices.",
    },
    {
        "paper": "Fetzer (2019)",
        "wave": "2015-2019",
        "tabla": "Table 1, Panel A, Col 1",
        "G": "id (constituency)", "T": "ryr (relative year)", "D": "temp (austerity exposure)", "Y": "pct_votes_UKIP",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 55.1% negative weights (483 pos / 592 neg). Austerity and rise of UKIP. District-level only (UKHLS not included).",
    },
    {
        "paper": "Kaur (2019)",
        "wave": "2015-2019",
        "tabla": "Table 1, Col 1",
        "G": "dist (district)", "T": "year", "D": "amons80 (monsoon rainfall shock)", "Y": "lwage",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. 1.4% negative weights (1708 pos / 24 neg). Nominal wage rigidity in India. Rainfall shock binary treatment.",
    },
    {
        "paper": "Suarez Serrato and Zidar (2016)",
        "wave": "2015-2019",
        "tabla": "Table 4, Panel A, Col 1",
        "G": "fe_group", "T": "year", "D": "d_keeprate (tax rate change)", "Y": "E (establishment count)",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "panel_GTD.dta construido. Panel structure incompatible con weight decomposition (rc=402). State corporate tax incidence.",
    },

    # --- REPLICADO NEW (Pierce & Schott) ---
    {
        "paper": "Pierce and Schott (2016)",
        "wave": "2015-2019",
        "tabla": "Table 1",
        "G": "fam50 (industry family)", "T": "year (1990-2007)", "D": "s1999 (NTR gap), s1999_post = s1999 x post(>=2001)", "Y": "lemp (log employment, NBER-CES)",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "Pipeline completo. 5814 obs (323 industries x 18 anos). twowayfeweights feTR: 1344 pos / 917 neg (40.6% negative). Design: HAD. Y usa NBER-CES public data (naics5809.dta). Spec: areg lemp s1999_post d???? [aw=emp1990], a(fam50) cl(fam50) robust. PDF compilado.",
    },

    # --- RESTRINGIDO (Wave 2) ---
    {
        "paper": "Allcott, Collard-Wexler, O'Connell (2016)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "plant/state", "T": "year", "D": "electricity shortage", "Y": "revenue/productivity (TFPR)",
        "G_d": "NO", "T_d": "NO", "D_d": "SI", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "16 .dta auxiliares presentes (deflactores, concordancias NIC, IO tables, hydroplants). Falta el dataset principal del Annual Survey of Industries (ASI) de India. ASI microdata se compra al Ministry of Statistics and Programme Implementation (MoSPI) de India, ~$2000-5000 dependiendo del periodo. Requiere solicitud formal via NSSO/CSO.",
    },
    {
        "paper": "Atkin (2016)",
        "wave": "2015-2019",
        "tabla": "Main panels",
        "G": "municipality", "T": "year", "D": "tariff reduction (NAFTA)", "Y": "productivity growth",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "13 .dta auxiliares presentes (concordancias IMSS-census, datos municipales). Datos principales de productividad provienen del IMSS (Instituto Mexicano del Seguro Social) panel de empleadores. IMSS requiere convenio institucional con gobierno mexicano y aprobacion de comite de datos. No hay acceso publico a microdatos de IMSS.",
    },
    {
        "paper": "Besley, Folke, Persson, Rickne (2017)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "municipality/party", "T": "election year", "D": "gender quota adoption", "Y": "competence index",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Cero .dta. Solo 5 .do files y documentacion. Datos de registros administrativos suecos protegidos por Swedish Secrecy Act (Offentlighets- och sekretesslagen). Datos accesibles solo via Statistics Sweden (SCB). Investigadores extranjeros deben tramitar acceso remoto via MONA (Microdata Online Access), proceso ~6 meses, requiere afiliacion sueca o coautoria.",
    },
    {
        "paper": "Bloom, Brynjolfsson, Foster, Jarmin, Patnaik, Saporta-Eksten, Van Reenen (2019)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "plant/firm", "T": "year", "D": "management practices (structured)", "Y": "productivity",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Cero .dta. Solo README. Datos de MOPS (Management and Organizational Practices Survey) ligados a Census of Manufacturers y LBD del US Census Bureau. Acceso exclusivo via Census Research Data Centers (RDCs federales). Proceso: propuesta de proyecto, security clearance, ~6-12 meses de aprobacion. Datos agregados publicos en census.gov/mops pero insuficientes para replicacion.",
    },
    {
        "paper": "Brandt, Van Biesebroeck, Wang, Zhang (2017)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "firm", "T": "year", "D": "WTO tariff reduction", "Y": "TFP/employment",
        "G_d": "NO", "T_d": "NO", "D_d": "SI", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "5 .dta auxiliares presentes (deflactores, concordancias HS-CIC, variables FTC, proteccion comercial). Datos principales de Annual Survey of Above-scale Industrial Firms del NBS China (National Bureau of Statistics). Solo accesible por permiso del NBS o data sellers autorizados (ej. GTA). No hay acceso publico; datos frecuentemente vendidos ~$5000-15000 por decada.",
    },
    {
        "paper": "Di Maggio, Kermani, Keys, Piskorski, Ramcharan, Seru, Yao (2017)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "zip code/county", "T": "quarter", "D": "ARM share (rate reset exposure)", "Y": "mortgage refinancing/consumption",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Cero .dta. Solo 1 .do file y documentacion. Datos de tres fuentes propietarias: (1) BlackBox Logic (ahora ICE Mortgage Technology) loan-level data ~$50K+/ano; (2) CoreLogic LLMA (loan-level servicing) ~$25K+/ano; (3) HMDA (Home Mortgage Disclosure Act, publico para agregados pero microdata via CFPB). La combinacion de las tres fuentes hace la replicacion economicamente inviable.",
    },
    {
        "paper": "Diamond, McQuade, Qian (2019)",
        "wave": "2015-2019",
        "tabla": "Tables 4-7",
        "G": "neighborhood/tract", "T": "year", "D": "LIHTC project placement", "Y": "house prices/demographics",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Restringido",
        "nota": "5 .dta presentes (ltdb_std_all_2010_adjusted.dta, usa_00012.dta, clean_combined.dta, block group demographics). Datos censales y LIHTC publicos. PERO el panel principal usa datos de Infutor Data Solutions (address-level consumer data). Infutor (ahora Verisk) vende datos de cambio de domicilio bajo licencia comercial; no disponible para investigadores sin contrato. Completa estructura de .do files presente (20+ scripts).",
    },
    {
        "paper": "Fuest, Peichl, Siegloch (2018)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "municipality/firm", "T": "year", "D": "local business tax rate change", "Y": "wages",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Cero .dta. Solo 11 .do files (master.do + data prep + estimation). Datos de registros de seguridad social alemanes del IAB (Institute for Employment Research). Acceso solo via FDZ (Forschungsdatenzentrum) del IAB tras solicitud formal. Datos procesados en servidores del FDZ; no se pueden exportar microdatos. Requiere remote execution via JoSuA (Job Submission Application). Proteccion de datos alemana (BDSG) aplica.",
    },
    {
        "paper": "Huber (2018)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "firm/county", "T": "year", "D": "bank lending shock (Commerzbank)", "Y": "employment/investment",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Cero .dta. Solo 1 .do file. Datos de dos fuentes propietarias: (1) Creditreform (agencia de credito alemana, datos de relaciones bancarias de firmas); contactar info@creditreform.com. (2) Bureau van Dijk/Dafne (datos financieros de firmas); acceso via BvD/Moody's con licencia ~$10K-30K/ano. Datos no disponibles publicamente.",
    },
    {
        "paper": "Naritomi (2019)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "firm", "T": "month", "D": "NFP adoption (lottery incentive)", "Y": "tax revenue",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "5 .dta auxiliares presentes (Google searches, PAC data, PIB IBGE, tax_bacen). Datos principales de la Secretaria de Finanzas del Estado de Sao Paulo (SEFAZ/SP). Requiere autorizacion del SEFAZ: (1) contactar a la autora para autorizacion y acceso en LSE, o (2) firmar memorandum de entendimiento (MoU) directamente con SEFAZ/SP. Datos de firmas brasilenas protegidos por sigilo fiscal.",
    },

    # --- OTROS Wave 2 (no TWFE, no folder, etc.) ---
    {
        "paper": "Dell (2015)",
        "wave": "2015-2019",
        "tabla": "Tables 1, 2B, 3B",
        "G": "municipality", "T": "election period", "D": "PAN victory (RDD)", "Y": "homicides/drug trade",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "Datos completos disponibles. Replicacion completa con PDF. Diseno RDD (regression discontinuity), NO es TWFE estandar con panel de unidades fijas. No aplica twowayfeweights. Drug-specific columns redacted por seguridad.",
    },
    {
        "paper": "Handley and Limao (2017)",
        "wave": "2015-2019",
        "tabla": "6 main + 9 appendix tables",
        "G": "product/firm", "T": "year", "D": "trade policy uncertainty (WTO accession)", "Y": "imports/prices",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "Datos completos, 17 tablas replicadas con PDF. NO es TWFE estandar: cross-sectional, long-difference, y multi-way FE designs. No aplica twowayfeweights.",
    },
    {
        "paper": "Hershbein and Kahn (2018)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "MSA", "T": "year", "D": "Bartik shock x post", "Y": "skill requirements (job postings)",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "Datos completos. NO es TWFE estandar: cross-sectional Bartik design sin FE de unidad. No aplica twowayfeweights.",
    },
    {
        "paper": "Munshi and Rosenzweig (2016)",
        "wave": "2015-2019",
        "tabla": "Tables 6, 8a",
        "G": "caste/jati", "T": "survey round", "D": "caste network strength", "Y": "mobility/schooling",
        "G_d": "SI", "T_d": "SI", "D_d": "SI", "Y_d": "SI",
        "estado": "Replicado",
        "nota": "Datos completos, replicacion parcial con PDF. Cross-sectional caste networks, no TWFE estandar con panel. cgmwildboot issue en algunas tablas.",
    },
    {
        "paper": "Hoynes, Miller, Simon (2016)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "state/demographic group", "T": "year", "D": "EITC expansion (phase-in rate)", "Y": "infant health (birth weight, prematurity)",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Restringido",
        "nota": "Sin carpeta en el repositorio (no descargado). Datos de Vital Statistics Natality microdata del NCHS (National Center for Health Statistics). La version publica (disponible en NBER) tiene variables geograficas limitadas (estado). La version restricted-use tiene condado y requiere acceso al NCHS RDC via propuesta formal. Datos confidenciales bajo 42 USC Section 242m(d).",
    },
    {
        "paper": "Dix-Carneiro and Kovak (2017)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "microregion", "T": "year", "D": "regional tariff change (trade liberalization)", "Y": "earnings/employment",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Falta data",
        "nota": "Sin carpeta en el repositorio (no descargado). Datos del Censo Demografico y RAIS (Relacao Anual de Informacoes Sociais) de Brasil. RAIS es semi-publico: disponible para investigadores brasileros via MTE (Ministerio do Trabalho). Acceso internacional requiere convenio con MTE. Censo publico via IBGE pero nivel microregion requiere microdatos.",
    },
    {
        "paper": "Monte, Redding, Rossi-Hansberg (2018)",
        "wave": "2015-2019",
        "tabla": "Main tables",
        "G": "commuting zone/county", "T": "year", "D": "trade/productivity shock", "Y": "commuting flows/wages",
        "G_d": "NO", "T_d": "NO", "D_d": "NO", "Y_d": "NO",
        "estado": "Falta data",
        "nota": "Sin carpeta en el repositorio (no descargado). Datos del LEHD (Longitudinal Employer-Household Dynamics) del US Census Bureau, acceso solo via Census RDCs. Tambien usa American Community Survey (publico) y County Business Patterns (publico).",
    },
]

# ============================================================================
# WRITE DATA ROWS
# ============================================================================

# Sort: first by wave, then Replicado first, then alphabetically
def sort_key(p):
    wave_order = 0 if p["wave"] == "2010-2012" else 1
    estado_order = {"Replicado": 0, "Restringido": 1, "Falta data": 2}
    return (wave_order, estado_order.get(p["estado"], 3), p["paper"])

papers.sort(key=sort_key)

for i, p in enumerate(papers, 1):
    row = i + 1  # header is row 1
    estado = p["estado"]

    # Pick fill color
    if estado == "Replicado":
        fill = GREEN
    elif estado == "Falta data":
        fill = YELLOW
    else:
        fill = RED

    values = [
        i,
        p["paper"],
        "AER",
        p["wave"],
        p["tabla"],
        p["G"],
        p["T"],
        p["D"],
        p["Y"],
        p["G_d"],
        p["T_d"],
        p["D_d"],
        p["Y_d"],
        estado,
        p["nota"],
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Apply fill to Estado column and the whole row
        if col == 14:  # Estado column
            cell.font = BOLD_FONT
            cell.fill = fill
        # Color the availability columns
        if col in (10, 11, 12, 13):
            cell.alignment = Alignment(horizontal="center", vertical="top")
            if val == "SI":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif val == "NO":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif "PARCIAL" in str(val):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# ============================================================================
# COLUMN WIDTHS
# ============================================================================
col_widths = {
    1: 4,    # #
    2: 42,   # Paper
    3: 6,    # Journal
    4: 12,   # Wave
    5: 28,   # Tabla
    6: 30,   # G
    7: 25,   # T
    8: 40,   # D
    9: 35,   # Y
    10: 7,   # G disp
    11: 7,   # T disp
    12: 7,   # D disp
    13: 7,   # Y disp
    14: 13,  # Estado
    15: 80,  # Nota
}

for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Row height for header
ws.row_dimensions[1].height = 30

# ============================================================================
# SUMMARY SHEET
# ============================================================================
ws2 = wb.create_sheet("Resumen")

summary_data = [
    ["RESUMEN ESTADO PAPERS - TWFE SURVEY", "", ""],
    ["", "", ""],
    ["Categoria", "Cantidad", "Porcentaje"],
    ["Replicado (panel_GTD.dta o analisis completo)", 0, ""],
    ["Restringido (datos de acceso restringido)", 0, ""],
    ["Falta data (sin datos en el package)", 0, ""],
    ["", "", ""],
    ["TOTAL", 0, "100%"],
]

# Count
n_rep = sum(1 for p in papers if p["estado"] == "Replicado")
n_res = sum(1 for p in papers if p["estado"] == "Restringido")
n_fal = sum(1 for p in papers if p["estado"] == "Falta data")
n_total = len(papers)

summary_data[3][1] = n_rep
summary_data[3][2] = f"{n_rep/n_total*100:.1f}%"
summary_data[4][1] = n_res
summary_data[4][2] = f"{n_res/n_total*100:.1f}%"
summary_data[5][1] = n_fal
summary_data[5][2] = f"{n_fal/n_total*100:.1f}%"
summary_data[7][1] = n_total

for r, row_data in enumerate(summary_data, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = Font(name="Calibri", size=14, bold=True)
        elif r == 3:
            cell.font = BOLD_FONT
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        elif r == 4:
            cell.fill = GREEN
        elif r == 5:
            cell.fill = RED
        elif r == 6:
            cell.fill = YELLOW
        elif r == 8:
            cell.font = BOLD_FONT

ws2.column_dimensions["A"].width = 50
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12

# ============================================================================
# SAVE
# ============================================================================
outpath = r"C:\Users\Usuario\Documents\GitHub\twfe_survey\reports\estado_papers_final.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
print(f"Total papers: {n_total}")
print(f"  Replicado: {n_rep}")
print(f"  Restringido: {n_res}")
print(f"  Falta data: {n_fal}")
