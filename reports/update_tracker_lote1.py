"""
Update tracker_master_consolidado.xlsx for Lote 1:
- 3 completed papers: Handley & Limao, Anderson & Sallee, Bloom et al.
- 4 inviable papers: Bajari, Duggan, Dahl & Lochner, Hershbein & Kahn
"""
import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(
    'C:/Users/Usuario/Documents/GitHub/twfe_survey/reports/tracker_master_consolidado.xlsx'
)
ws = wb.active

# Colors matching existing convention
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Row 18: Anderson (2011)
ws['G18'] = 'Replicado (parcial)'
ws['G18'].fill = blue_fill
ws['H18'] = 'Parcial: cafe_compliance.dta disponible. Tables 1, 2, 8 + Figures 1-5 replicados. transactions1.dta propietaria (Tables 3-7 no replicables).'

# Row 19: Bajari et al. (2012) — already Sin data, update notes
ws['G19'] = 'Sin data'
ws['G19'].fill = red_fill
ws['H19'] = 'INVIABLE. DataQuick housing data propietaria. Datos no incluidos en paquete de replicacion.'

# Row 22: Bloom et al. (2012)
ws['G22'] = 'Replicado (parcial)'
ws['G22'].fill = blue_fill
ws['H22'] = 'Parcial: European data disponible (replicate.dta). Tables 6, C2, A5, A6 replicados. UK census data (Tables 1-5, C1) requiere acceso VML ONS London.'

# Row 25: Dahl and Lochner (2012)
ws['G25'] = 'Sin data'
ws['G25'].fill = red_fill
ws['H25'] = 'INVIABLE. state var=0 (datos reales ausentes). Requiere TAXSIM externo + datos IRS restringidos. No replicable.'

# Row 27: Duggan et al. (2011)
ws['G27'] = 'Sin data'
ws['G27'].fill = red_fill
ws['H27'] = 'INVIABLE. IMS Health data (ims0106data2.dta) propietaria. No incluida en paquete de replicacion.'

# Row 51: Handley and Limao (2017)
ws['G51'] = 'Replicado'
ws['G51'].fill = blue_fill
ws['H51'] = 'Full replication. Tables 1-6, A1-A9, Figures 2-4 replicados. 6 main tables + 9 appendix tables + 8 figuras.'

# Row 56: Hershbein and Kahn (2018)
ws['G56'] = 'Sin data'
ws['G56'].fill = red_fill
ws['H56'] = 'INVIABLE. Burning Glass Technologies data propietaria. Datos no incluidos en paquete de replicacion.'

wb.save('C:/Users/Usuario/Documents/GitHub/twfe_survey/reports/tracker_master_consolidado.xlsx')
print("=== Tracker updated successfully ===")
print("Updated rows: 18 (Anderson), 19 (Bajari), 22 (Bloom), 25 (Dahl), 27 (Duggan), 51 (Handley), 56 (Hershbein)")
