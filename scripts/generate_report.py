"""
Generate AER 2010-2012 Feasibility Report Excel file.
Based on de Chaisemartin & D'Haultfoeuille (2020) web appendix Section 6.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Feasibility Report"

# ── Styles ──────────────────────────────────────────────────────────────────
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ── Headers ─────────────────────────────────────────────────────────────────
headers = [
    "Paper#", "Autores", "Titulo Completo", "AER Vol(Num), Paginas",
    "Tabla/Figura Objetivo", "Tipo Regresion", "Diseno",
    "Status", "Carpeta Disponible", "Archivos Clave",
    "Num Archivos", "Tamano Total", "Software Requerido",
    "Datos Faltantes / Notas", "Contacto Autor Principal", "Institucion"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin_border

# ── Column widths ───────────────────────────────────────────────────────────
widths = [7, 22, 45, 20, 25, 22, 10, 14, 28, 35, 10, 12, 14, 40, 30, 25]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Paper data ──────────────────────────────────────────────────────────────
# Format: (paper#, authors, title, vol_pages, target_table, regression_type, design,
#          status, folder, key_files, n_files, total_size, software, notes, contact, institution)

papers = [
    (1, "Chandra, Gruber, McKnight",
     "Patient Cost-Sharing and Hospitalization Offsets in the Elderly",
     "AER 100(1), 193-213",
     "Tables 2 & 3, first line",
     "Reg 1 (FE), plan & month FE",
     "Sharp",
     "ROJO",
     "Chandra et al. (2010)",
     "31 .do files, NO .dta datasets",
     31, "0.6 MB", "Stata",
     "SOLO CODIGO. Datos Medicare confidenciales. Requiere acceso a CMS/NBER. No hay .dta en el paquete.",
     "Amitabh Chandra: achandra@hks.harvard.edu",
     "Harvard Kennedy School"),

    (2, "Duggan, Morton",
     "The Effect of Medicare Part D on Pharmaceutical Prices and Utilization",
     "AER 100(1), 590-607",
     "Tables 2 & 3",
     "Reg 2 (FD) con controles",
     "Sharp",
     "AMARILLO",
     "Duggan and Morton (2010)",
     ".dta: mcd2003us, meps02/03person, meps02/03share, merge0203AA, otc0305, otc05info, usp548final, miscvars1; .do: regs-partd-final, otc0305, meps02/03share; .dct: h67a,h70,h77a,h79",
     26, "12.2 MB", "Stata",
     "Datos disponibles pero .dct sugiere que algunos datos raw (MEPS) deben descargarse de AHRQ. Verificar si meps*.dta son suficientes.",
     "Mark Duggan: mduggan@stanford.edu",
     "Stanford University"),

    (3, "Aizer",
     "The Gender Wage Gap and Domestic Violence",
     "AER 100(4), 1847-1859",
     "Table 2",
     "Reg 1 (FE) three-way + controles",
     "Fuzzy",
     "VERDE",
     "Aizer (2010)",
     ".dta: census (82MB), census2000 (89MB), gamma_linear, bartik03_aer, cadv_2003, clinicall, demogb03, drugs, immigration, shelters, etc.; .do: aer_analysis, aer_bartik, aer_census1990/2000, makeall03_aer",
     58, "175.4 MB", "Stata",
     "Datos completos. Incluye census 1990/2000, datos de violencia, industria. Listo para replicar.",
     "Anna Aizer: anna_aizer@brown.edu",
     "Brown University"),

    (4, "Algan, Cahuc",
     "Inherited Trust and Growth",
     "AER 100(5), 2060-2092",
     "Figure 4",
     "Reg 2 (FD)",
     "Sharp",
     "VERDE",
     "Algan and Cahuc (2010)",
     ".dta: AER_MACRO (37KB), AER_MICRO (16MB); .do: README_PROGRAM, run_twowayfe; .tex: tables output",
     12, "17.2 MB", "Stata",
     "Datos completos. Incluye macro y micro datasets. Ya tiene run_twowayfe.do con resultados.",
     "Yann Algan: yann.algan@sciencespo.fr",
     "Sciences Po Paris"),

    (5, "Ellul, Pagano, Panunzi",
     "Inheritance Law and Investment in Family Firms",
     "AER 100(5), 2414-2450",
     "Table 7",
     "Reg 1 (FE) + 3 treatments + controles",
     "Sharp",
     "ROJO",
     "Ellul et al. (2010)",
     ".do: Inheritance_1, Inheritance_2; readme.pdf. NO .dta datasets",
     5, "1.0 MB", "Stata",
     "SOLO CODIGO. Datos de firmas familiares confidenciales. Requiere acceso a base de datos propietaria de firmas.",
     "Andrew Ellul: aellul@indiana.edu",
     "Indiana University (Kelley)"),

    (6, "Bustos",
     "Trade Liberalization, Exports, and Technology Upgrading: Evidence on the Impact of MERCOSUR on Argentinean Firms",
     "AER 101(1), 304-340",
     "Tables 3 to 12",
     "Reg 2 (FD) con controles",
     "Sharp",
     "ROJO",
     "Bustos (2011)",
     ".do: 20081226_Create_Data, 20081226_Create_Results, 20081226_Tables_1_and_10; readme. NO .dta",
     6, "0.9 MB", "Stata",
     "SOLO CODIGO. Datos de firmas argentinas (encuesta industrial) no incluidos. Requiere acceso a datos INdEC.",
     "Paula Bustos: paula.bustos@cemfi.es",
     "CEMFI Madrid"),

    (7, "Anderson, Sallee",
     "Using Loopholes to Reveal the Marginal Cost of Regulation: The Case of Fuel-Economy Standards",
     "AER 101(4), 1375-1409",
     "Table 5, Column 2",
     "Reg 1 (FE), state & month FE + controles",
     "Fuzzy",
     "VERDE",
     "Anderson and Sallee (2011)",
     ".dta: cafe_compliance (659KB); .do: cafe_compliance_behavior, cafe_compliance_costs; .xlsx: Calculations_Table_1/6/8; transactions .do/.txt files",
     21, "2.3 MB", "Stata + Excel",
     "Datos completos. Incluye CAFE compliance data y transactions. Output en .txt files.",
     "Soren Anderson: sta@msu.edu",
     "Michigan State University"),

    (8, "Bagwell, Staiger",
     "What Do Trade Negotiators Negotiate About? Empirical Evidence from the World Trade Organization",
     "AER 101(4), 1238-1273",
     "Table 3, OLS columns",
     "Reg 1 (FE) similar, 2 treatments",
     "Sharp",
     "VERDE",
     "Bagwell and Staiger (2011)",
     ".txt: AVEData (2.1MB), MainData (8.8MB), NonLinearData2 (1.7MB); .dta: NonLinearData; .R: Basic1, BasicPrice1, NormalChina, NTB1, etc.; .do: nonlinearcode, run_twowayfe",
     27, "14.6 MB", "R + Stata",
     "Datos completos. Principalmente en R scripts. Ya tiene run_twowayfe.do con resultados twowayfeweights.",
     "Kyle Bagwell: kbagwell@stanford.edu",
     "Stanford University"),

    (9, "Zhang, Zhu",
     "Group Size and Incentives to Contribute: A Natural Experiment at Chinese Wikipedia",
     "AER 101(4), 1601-1605",
     "Tables 3 & 4, Columns 4-6",
     "Reg 1 (FE) + controles",
     "Sharp",
     "VERDE",
     "Zhang and Zhu (2011)",
     ".dta: contridaily (129MB), contridaily_proc (286MB), daily_contribution (143MB), daily_newpost_addition (86MB), new, old, noncontentious, nonblocked, etc.; .do: main, genweek, appendix_*; run_twowayfe.do",
     38, "1033.8 MB", "Stata",
     "Datos completos (1GB+). Contribuciones diarias Wikipedia china. Ya tiene run_twowayfe.do.",
     "Xiaoquan (Michael) Zhang: michael.zhang@cornell.edu",
     "Cornell University"),

    (10, "Hotz, Xiao",
     "The Impact of Regulations on the Supply and Quality of Care in Child Care Markets",
     "AER 101(5), 1775-1805",
     "Table 7, Columns 4 & 5",
     "Reg 1 (FE), state & year FE + controles",
     "Sharp",
     "VERDE",
     "Hotz and Xiao (2011)",
     ".dta: Census (c1990_clean, c2000_clean, census1990, census2000), zipbundle weights, Regulation data (reg, chcare_aizer), Childcare_nonemployer; .do: analysis, matching programs",
     78, "203.8 MB", "Stata + SAS",
     "Datos completos. Census, regulacion, childcare. Programa de merge complejo pero todo incluido.",
     "V. Joseph Hotz: hotz@econ.duke.edu",
     "Duke University"),

    (11, "Mian, Sufi",
     "House Prices, Home Equity-Based Borrowing, and the US Household Leverage Crisis",
     "AER 101(5), 2132-2156",
     "Tables 2 & 3",
     "Reg 2 (FD) 2SLS + controles",
     "Sharp",
     "ROJO",
     "Mian and Sufi (2011)",
     ".do: MianSufi_AER_homeequity (43KB); readme.txt. NO .dta datasets",
     5, "0.9 MB", "Stata",
     "SOLO CODIGO. Datos de hogares (Equifax, LPS) son propietarios/confidenciales. No se pueden replicar sin licencia de datos.",
     "Atif Mian: atif@princeton.edu",
     "Princeton University"),

    (12, "Wang",
     "State Misallocation and Housing Prices: Theory and Evidence from China",
     "AER 101(5), 2081-2107",
     "Table 5, Panel A",
     "Reg 1 (FE) similar + controles",
     "Sharp",
     "VERDE",
     "Wang (2011)",
     ".dta: data_aersubmit (7.5MB); .do: aer_wang_misallocation_tables, aer_wang_misallocation_summary; readme.pdf",
     6, "8.2 MB", "Stata",
     "Datos completos. Un solo dataset con todo. Listo para replicar.",
     "Shaoda Wang: shaodaw@uchicago.edu",
     "University of Chicago"),

    (13, "Duranton, Turner",
     "The Fundamental Law of Road Congestion: Evidence from US Cities",
     "AER 101(6), 2616-2652",
     "Table 5",
     "Reg 2 (FD)",
     "Sharp",
     "VERDE",
     "Duranton and Turner (2011)",
     ".dta: Duranton_Turner_AER_2010 (322KB); .do: Duranton_Turner_AER_2010 (120KB); ReadMe-data.txt",
     5, "1.8 MB", "Stata",
     "Datos completos. Un dataset y un .do file completo. Muy limpio para replicar.",
     "Gilles Duranton: duranton@wharton.upenn.edu",
     "University of Pennsylvania (Wharton)"),

    (14, "Acemoglu, Cantoni, Johnson, Robinson",
     "The Consequences of Radical Reform: The French Revolution",
     "AER 101(7), 3286-3307",
     "Table 3",
     "Reg 1 (FE), polity & time FE",
     "Sharp",
     "VERDE",
     "Acemoglu et al. (2011)",
     ".dta: 20100816_replication_dataset (85KB), _dataset_t5 (20KB); .do: 20100816_replication10; readme.pdf",
     6, "0.9 MB", "Stata",
     "Datos completos. Dataset de replicacion con readme claro. Listo.",
     "Daron Acemoglu: daron@mit.edu",
     "MIT"),

    (15, "Baum-Snow, Lutz",
     "School Desegregation, School Choice, and Changes in Residential Location Patterns by Race",
     "AER 101(7), 3019-3046",
     "Tables 2 to 6",
     "Reg 1 (FE) + controles",
     "Sharp",
     "VERDE",
     "Baum-SnowandLutz(2011)",
     ".dta: tractpanx (85MB), cntypan (5MB), districtpan, msapan, GIS data (panel-seggis-all 39MB), census tracts 60-90, district 70-90; .do: tables2to5, table6, figure*, counterfact, panel-censusx, build-districts",
     120, "908.4 MB", "Stata",
     "Datos completos (908MB). Census tracts, GIS, district data. Paquete grande y complejo pero todo incluido.",
     "Nathaniel Baum-Snow: baum-snow@utoronto.ca",
     "University of Toronto"),

    (16, "Dinkelman",
     "The Effects of Rural Electrification on Employment: New Evidence from South Africa",
     "AER 101(7), 3078-3108",
     "Tables 4,5 Cols 5-8; Table 8 Cols 3-4; Table 9 Col 2; Table 10 Cols 2,4,6",
     "Reg 2 (FD) 2SLS + controles",
     "Sharp",
     "VERDE",
     "Dinkelman (2011)",
     ".dta: censusmicrodata (19MB), hhsurveydata (1.7MB), kidsunder9_96, matched_censusdata, placebodata, lfs02_migrantdata, census_comm_indiv; .do: master, mainanalysis_communitydata, other, supplanalysis_*; .ado: x_gmm_td, x_ols_td",
     18, "27.9 MB", "Stata",
     "Datos completos. Census micro data + household survey. Custom .ado incluidos.",
     "Taryn Dinkelman: taryn.dinkelman@nd.edu",
     "University of Notre Dame"),

    (17, "Enikolopov, Petrova, Zhuravskaya",
     "Media and Political Persuasion: Evidence from Russia",
     "AER 101(7), 3253-3285",
     "Table 3",
     "Reg 1 (FE), subregion & election FE",
     "Fuzzy",
     "VERDE",
     "Enikolopov et al. (2011)",
     ".dta: NTV_Aggregate_Data (775KB), NTV_Individual_Data (1.2MB); .do: Aggregate_level_results_Replication, Individual_level_results_Replication; README.pdf",
     10, "2.8 MB", "Stata",
     "Datos completos. Aggregate + Individual level. Clean replication package.",
     "Ruben Enikolopov: ruben.enikolopov@upf.edu",
     "Universitat Pompeu Fabra / NES"),

    (18, "Fang, Gavazza",
     "Dynamic Inefficiencies in an Employment-Based Health Insurance System: Theory and Evidence",
     "AER 101(7), 3047-3077",
     "Tables 2, 3, 5, 6 Column 3",
     "Reg 2 (FD) 2SLS + controles",
     "Fuzzy",
     "VERDE",
     "Fang and Gavazza",
     ".dta: BHPS waves e-n (indresp/hhresp 13-38MB each), MEPS h12-h70 (40-91MB each), rndhrs_h (342MB), statesector (51MB), dyn_st_89_98s1, ads-laws; .do: tables2_3_5, tables2_3_5_col4, table4, table6, table7, firms4",
     60, "1205.6 MB", "Stata",
     "Datos completos (1.2GB). BHPS + MEPS + HRS data. Paquete grande pero todo incluido.",
     "Hanming Fang: hanming.fang@econ.upenn.edu",
     "University of Pennsylvania"),

    (19, "Gentzkow, Shapiro, Sinkinson",
     "The Effect of Newspaper Entry and Exit on Electoral Politics",
     "AER 101(7), 2980-3018",
     "Tables 2 & 3",
     "Reg 2 (FD) + controles",
     "Sharp",
     "VERDE",
     "Gentzkow et al. (2011)",
     ".dta: newspapers_constant, newspapers_yearly, voting_cnty_clean (31MB), voting_district_clean, lifecycle, nparchive, endorse, suffrage, journalists; .do: tables, figures, text; .m: simulate, pakes_ostrovsky_berry; .ado: custom commands",
     82, "70.8 MB", "Stata + MATLAB",
     "Datos completos. Newspaper + voting data. Incluye simulaciones MATLAB. Replication package muy completo.",
     "Matthew Gentzkow: gentzkow@stanford.edu",
     "Stanford University"),

    (20, "Bloom, Sadun, Van Reenen",
     "Americans Do IT Better: US Multinationals and the Productivity Miracle",
     "AER 102(1), 167-201",
     "Table 2, Columns 6-8",
     "Reg 1 (FE) + 3 treatments + controles",
     "Sharp",
     "AMARILLO",
     "Bloom et al. (2012)",
     "European-Results/: replicate.dta (1.2MB), Figures.dta, sampling_final_TableA5.dta; .do: Table6_C2, TableA5, TableA6, Figures; UK-Data-Preparation/: 1_ARD, 1_BSCI, 1_FAR, 1_QICE (data build); UK-Results/: ADIB_Alltables",
     37, "5.3 MB", "Stata",
     "European data disponible (Table 6, A5, A6). UK data (ARD, BSCI, QICE) requiere acceso a ONS Secure Lab. Table 2 cols 6-8 puede requerir UK data.",
     "Nick Bloom: nbloom@stanford.edu",
     "Stanford University"),

    (21, "Simcoe",
     "Standard Setting Committees: Consensus Governance for Shared Technology Platforms",
     "AER 102(1), 305-336",
     "Table 4, Columns 1-3",
     "Reg 1 (FE) similar + controles",
     "Sharp",
     "VERDE",
     "Simcoe (2012)",
     ".dta: idLevel (4.8MB), idVersionLevel (4.8MB), bigFirmDummies; .out: email_panel, email_panel_gt4, email_panel_replies; .do: analysis, olsMainRegs, olsRobust, olsInteractions, matching, figures, citeCounts, endogSwitch; .log: full outputs",
     64, "15.5 MB", "Stata",
     "Datos completos. IETF standards data. Muchos .do y .log con resultados previos.",
     "Timothy Simcoe: tsimcoe@bu.edu",
     "Boston University"),

    (22, "Moser, Voena",
     "Compulsory Licensing: Evidence from the Trading with the Enemy Act",
     "AER 102(1), 396-427",
     "Table 2",
     "Reg 1 (FE) + controles",
     "Sharp",
     "VERDE",
     "Moser and Voena (2012)",
     ".dta: chem_patents_maindataset (23MB), chem_patents_primaryclassesdataset (14MB), chem_patents_indigodataset (1.9MB), dupont_data (61MB), fig1/5/10, table1; .do: code_for_replication, code_for_replication_revised_2017; compulsory.xlsx",
     14, "228.8 MB", "Stata",
     "Datos completos. Patent data + DuPont data. Incluye version revisada 2017 del codigo.",
     "Petra Moser: pmoser@stern.nyu.edu",
     "NYU Stern"),

    (23, "Forman, Goldfarb, Greenstein",
     "The Internet and Local Wages: A Puzzle",
     "AER 102(1), 556-575",
     "Tables 2 & 4",
     "Reg 2 (FD) + controles",
     "Fuzzy",
     "VERDE",
     "Forman et al. (2012)",
     ".dta: countygrowth (826KB), countyyear (10.8MB); .do: tables (26KB); readme.pdf",
     6, "12.5 MB", "Stata",
     "Datos completos. County-level data. Package simple y limpio.",
     "Chris Forman: chris.forman@cornell.edu",
     "Cornell University"),

    (24, "Besley, Mueller",
     "Estimating the Peace Dividend: The Impact of Violence on House Prices in Northern Ireland",
     "AER 102(2), 810-833",
     "Table 1, Columns 3 & 5-7",
     "Reg 1 (FE), region & time FE",
     "Sharp",
     "VERDE",
     "Besley and Mueller",
     ".dta: maindata (367KB), tourismandkillings (12KB), israelpalestinequart (2.4KB); .txt: israelquart, totaldeathsall; .do: table-1, table3, table-4; pv5generation; SMOOTHSIMPLE; readme.pdf",
     13, "1.6 MB", "Stata",
     "Datos completos. Northern Ireland house prices + violence data. Incluye Israel-Palestine data adicional.",
     "Timothy Besley: t.besley@lse.ac.uk",
     "London School of Economics"),

    (25, "Dafny, Duggan, Ramanarayanan",
     "Paying a Premium on Your Premium? Consolidation in the US Health Insurance Industry",
     "AER 102(2), 1161-1185",
     "Table 3",
     "Reg 1 (FE) + controles",
     "Sharp",
     "ROJO",
     "Dafny et al. (2012)",
     ".do: dataset_create, Tables2_5, OA-Table-3, Table-6; Read-Me.docx/pdf. NO .dta datasets",
     8, "1.2 MB", "Stata",
     "SOLO CODIGO. Datos de seguros de salud (Large Group premiums) son propietarios. Imposible replicar sin licencia.",
     "Leemore Dafny: l-dafny@kellogg.northwestern.edu",
     "Harvard Business School"),

    (26, "Hornbeck",
     "The Enduring Impact of the American Dust Bowl: Short- and Long-Run Adjustments to Environmental Catastrophe",
     "AER 102(4), 1477-1507",
     "Table 2",
     "Reg 1 (FE) similar + 2 treatments + controles",
     "Fuzzy",
     "VERDE",
     "Hornbeck (2012)",
     ".dta: DustBowl_All_base1910 (31MB), Generate-Data/ (22 ICPSR .dta, erosion/centroids/woodland .txt, Export_*1910.txt), farmval, icpsr_fips, migclim; .do: Analyze_DustBowl (150KB), Generate_DustBowl (210KB); .xlsx: Create-Tables-and-Figures; .ado: x_ols",
     74, "203.5 MB", "Stata",
     "Datos completos. ICPSR census + GIS erosion data. Incluye tablas generadas en .docx. Package muy completo.",
     "Richard Hornbeck: rhornbeck@uchicago.edu",
     "University of Chicago"),

    (27, "Bajari, Fruehwirth, Kim, Timmins",
     "A Rational Expectations Approach to Hedonic Price Regressions with Time-Varying Unobserved Product Attributes: The Price of Pollution",
     "AER 102(5), 1898-1926",
     "Table 5",
     "Reg 2 (FD) similar + controles",
     "Sharp",
     "VERDE",
     "Bajari et al. (2012)",
     ".txt: data2sales (5.6MB), data3sales (1.3MB); .m: bckt_hedonic_2snls, secondstep*_obj; .do: crosssection, genfixedeffects, runfixedeffects, genyeardummy; README.docx/pdf",
     15, "7.9 MB", "MATLAB + Stata",
     "Datos completos. Housing sales data. Requiere MATLAB para main hedonic estimation, Stata para FE.",
     "Patrick Bajari: bajari@uw.edu",
     "University of Washington / Amazon"),

    (28, "Dahl, Lochner",
     "The Impact of Family Income on Child Achievement: Evidence from the Earned Income Tax Credit",
     "AER 102(5), 1927-1956",
     "Table 3",
     "Reg 2 (FD) 2SLS sin year FE",
     "Sharp",
     "VERDE",
     "Dahl and Lochner (2012)",
     ".dta: main (32.5MB); .do: main, makevars, regressions, taxsim-eitc, merge-school-welfare; .csv: welfare_dat; Read_me.pdf",
     10, "33.5 MB", "Stata",
     "Datos completos. NLSY child data + EITC. Dataset main tiene todo lo necesario.",
     "Gordon Dahl: gdahl@ucsd.edu",
     "UC San Diego"),

    (29, "Imberman, Kugler, Sacerdote",
     "Katrina's Children: Evidence on the Structure of Peer Effects from Hurricane Evacuees",
     "AER 102(5), 2048-2082",
     "Tables 3-6",
     "Reg 1 (FE) three-way + controles",
     "Fuzzy",
     "ROJO",
     "Imberman et al. (2012)",
     "83 .do files (quartile_analysis_*, katrina_*, merge_*, school_level_*, etc.); replication-readme.docx/pdf. NO .dta datasets",
     83, "2.9 MB", "Stata",
     "SOLO CODIGO (83 .do files). Datos de estudiantes (Houston/Louisiana ISD) son confidenciales. Requiere acuerdo con distrito escolar.",
     "Scott Imberman: imberman@msu.edu",
     "Michigan State University"),

    (30, "Chaney, Sraer, Thesmar",
     "The Collateral Channel: How Real Estate Shocks Affect Corporate Investment",
     "AER 102(6), 2381-2409",
     "Table 5",
     "Reg 1 (FE) + controles",
     "Sharp",
     "VERDE",
     "Chaney et al. (2012)",
     ".dta: headquarter_1997 (1.6MB), headquarter_2000 (2.9MB), adj_price, elasticity, msa_fips; .txt: housing_data, housing_data_MSA, interest, population, us_cpi_adj; .csv: zip2 (15.7MB); .do: master, reg, bootstrap, construc, first_stage, diff_in_diff, bubble, adj_price",
     23, "23.5 MB", "Stata",
     "Datos completos. Compustat + housing data. Master .do file organiza todo.",
     "Thomas Chaney: thomas.chaney@sciencespo.fr",
     "Sciences Po Paris"),

    (31, "Aaronson, Agarwal, French",
     "The Spending and Debt Response to Minimum Wage Hikes",
     "AER 102(7), 3111-3139",
     "Tables 1, 2, & 5",
     "Reg 1 (FE), household & time FE",
     "Sharp",
     "AMARILLO",
     "Aaronson et al. (2012)",
     ".dta.gz: ces_int_82_08 (91MB), ogr_select (277MB), rep_sipp1 (228MB); .zip: AER_FINAL_PROGRAMS; .gau: GAUSS scripts (minwageDP_*, initializations, etc.); .sas: SCF programs; .do: SS_*, scf, pooled_regs_fe; C/: DLL + C++ source; Version-Table.xlsx",
     53, "605.4 MB", "GAUSS + Stata + SAS + C++",
     "Datos en .gz comprimidos (debe descomprimir). Requiere GAUSS (software comercial) para structural model. .do files para tablas descriptivas. Software mix complejo.",
     "Daniel Aaronson: daniel.aaronson@chi.frb.org",
     "Federal Reserve Bank of Chicago"),

    (32, "Brambilla, Lederman, Porto",
     "Exports, Export Destinations, and Skills",
     "AER 102(7), 3406-3438",
     "Table 5",
     "Reg 1 (FE), firm & industry*period FE + controles",
     "Sharp",
     "VERDE",
     "Brambilla et al. (2012)",
     ".dta: comtradehs6 (77MB), countryclass, exports, ladder_hs6, pwt, usitc_tc; .do: data, table01-table12, _boot; readme.pdf",
     23, "77.6 MB", "Stata",
     "Datos completos. COMTRADE + firm data Argentina. 12 table .do files. Package completo.",
     "Irene Brambilla: irene.brambilla@econo.unlp.edu.ar",
     "Universidad Nacional de La Plata"),

    (33, "Faye, Niehaus",
     "Political Aid Cycles",
     "AER 102(7), 3516-3530",
     "Table 3 Cols 4-5; Tables 4 & 5",
     "Reg 1 (FE), donor*receiver FE",
     "Sharp",
     "VERDE",
     "Faye and Niehaus (2012)",
     ".dta: multiple baseline/estimation/final datasets (13-143MB each), CRS data, DPI elections, ICRG, NED grants, Pew, UN votes, WDI; .r: analysis_allR_v3, analysis_allR_sensitivity_v3, reg_helper_functions, lib_*; .do: cleaning code (oda_*)",
     71, "1231.8 MB", "R + Stata",
     "Datos completos (1.2GB). ODA + elections + outside data. Analisis en R, limpieza en Stata.",
     "Michael Faye: michael.faye@yale.edu",
     "Yale University / GiveDirectly"),
]

# ── Write data rows ─────────────────────────────────────────────────────────
status_fills = {"VERDE": green_fill, "AMARILLO": yellow_fill, "ROJO": red_fill}

for row_idx, p in enumerate(papers, 2):
    for col_idx, val in enumerate(p, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if col_idx == 8:  # Status column
            fill = status_fills.get(val)
            if fill:
                cell.fill = fill
                cell.font = Font(bold=True)

# ── Freeze panes ────────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Summary sheet ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resumen")

summary_data = [
    ["RESUMEN DE FACTIBILIDAD", ""],
    ["", ""],
    ["Total papers analizados:", 33],
    ["", ""],
    ["VERDE (datos completos, listo):", sum(1 for p in papers if p[7] == "VERDE")],
    ["AMARILLO (datos parciales):", sum(1 for p in papers if p[7] == "AMARILLO")],
    ["ROJO (datos faltantes/confidenciales):", sum(1 for p in papers if p[7] == "ROJO")],
    ["", ""],
    ["Papers VERDES:", ""],
]

green_papers = [f"  {p[0]}. {p[1]} ({p[3].split(',')[0]})" for p in papers if p[7] == "VERDE"]
yellow_papers = [f"  {p[0]}. {p[1]} ({p[3].split(',')[0]})" for p in papers if p[7] == "AMARILLO"]
red_papers = [f"  {p[0]}. {p[1]} ({p[3].split(',')[0]})" for p in papers if p[7] == "ROJO"]

for gp in green_papers:
    summary_data.append([gp, ""])
summary_data.append(["", ""])
summary_data.append(["Papers AMARILLOS:", ""])
for yp in yellow_papers:
    summary_data.append([yp, ""])
summary_data.append(["", ""])
summary_data.append(["Papers ROJOS:", ""])
for rp in red_papers:
    summary_data.append([rp, ""])

summary_data.extend([
    ["", ""],
    ["NOTAS:", ""],
    ["- ROJO = datos confidenciales/propietarios, requiere solicitud especial", ""],
    ["- AMARILLO = datos parciales o software especial requerido", ""],
    ["- VERDE = datos + codigo completos, listo para replicar con twowayfeweights", ""],
    ["", ""],
    ["Software requerido:", ""],
    ["- Stata: 33/33 papers", ""],
    ["- R: 3 papers (Bagwell & Staiger, Faye & Niehaus, + twowayfeweights)", ""],
    ["- MATLAB: 2 papers (Bajari et al., Gentzkow et al.)", ""],
    ["- GAUSS: 1 paper (Aaronson et al.)", ""],
    ["- SAS: 1 paper (Aaronson et al.)", ""],
    ["- Excel: 1 paper (Anderson & Sallee)", ""],
    ["", ""],
    ["Fuente: Web Appendix de 'Two-Way Fixed Effects Estimators", ""],
    ["         with Heterogeneous Treatment Effects'", ""],
    ["Autores: de Chaisemartin & D'Haultfoeuille (2020, AER)", ""],
    ["Seccion 6: Detailed Literature Review", ""],
])

for row_idx, (a, b) in enumerate(summary_data, 1):
    ws2.cell(row=row_idx, column=1, value=a)
    ws2.cell(row=row_idx, column=2, value=b)

ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
ws2.column_dimensions["A"].width = 60
ws2.column_dimensions["B"].width = 15

# Green/Yellow/Red formatting in summary
for row_idx, (a, b) in enumerate(summary_data, 1):
    if "VERDE" in str(a) and ":" in str(a) and "datos" not in str(a).lower():
        ws2.cell(row=row_idx, column=1).fill = green_fill
        ws2.cell(row=row_idx, column=1).font = Font(bold=True)
    elif "AMARILLO" in str(a) and ":" in str(a) and "datos" not in str(a).lower():
        ws2.cell(row=row_idx, column=1).fill = yellow_fill
        ws2.cell(row=row_idx, column=1).font = Font(bold=True)
    elif "ROJO" in str(a) and ":" in str(a) and "datos" not in str(a).lower():
        ws2.cell(row=row_idx, column=1).fill = red_fill
        ws2.cell(row=row_idx, column=1).font = Font(bold=True)

# ── Save ────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\Usuario\Documents\GitHub\papers_economic\AER_2011-2012_Feasibility_Report.xlsx"
wb.save(output_path)
print(f"Report saved to: {output_path}")
print(f"Total papers: {len(papers)}")
print(f"  VERDE:    {sum(1 for p in papers if p[7] == 'VERDE')}")
print(f"  AMARILLO: {sum(1 for p in papers if p[7] == 'AMARILLO')}")
print(f"  ROJO:     {sum(1 for p in papers if p[7] == 'ROJO')}")
