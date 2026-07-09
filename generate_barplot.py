"""
generate_barplot.py
Generate classification_barplot.png from summary_comparison_final_report.xlsx
Column: Final_classification  (Sheet: Summary)
"""

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from collections import Counter

# ── 1. Read data ──────────────────────────────────────────────────────────────
EXCEL = "/workspace/summary_comparison_final_report.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["Summary"]

headers = [c.value for c in ws[1]]
fc_col  = headers.index("Final_classification")

raw = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        val = row[fc_col]
        if val is None:
            val = "missing"
        raw.append(str(val).strip())

# ── 2. Normalise ──────────────────────────────────────────────────────────────
def normalise(v):
    v = v.strip()
    if v.upper() in {"OTHER", "OTHER "}:
        return "OTHER"
    return v

labels_raw = [normalise(v) for v in raw]

# ── 3. Count and order ───────────────────────────────────────────────────────
# Desired display order (estimable designs first, then edge-cases)
ORDER = ["CLA", "SAD", "HAD", "SFSD", "SSFSD",
         "OTHER", "WEIRD", "No time variable", "missing"]

counts = Counter(labels_raw)

# Add any unexpected labels at the end
for k in sorted(counts):
    if k not in ORDER:
        ORDER.append(k)

categories = [k for k in ORDER if k in counts]
values     = [counts[k] for k in categories]

# ── 4. Colours ───────────────────────────────────────────────────────────────
PALETTE = {
    "CLA"             : "#2196F3",   # blue
    "SAD"             : "#4CAF50",   # green
    "HAD"             : "#F44336",   # red
    "SFSD"            : "#FF9800",   # orange
    "SSFSD"           : "#FFB74D",   # light orange
    "SSD"             : "#9C27B0",   # purple (for future use)
    "OTHER"           : "#78909C",   # blue-grey
    "WEIRD"           : "#B0BEC5",   # lighter grey
    "No time variable": "#CFD8DC",   # very light grey
    "missing"         : "#ECEFF1",   # near-white grey
}
colors = [PALETTE.get(c, "#90A4AE") for c in categories]

# ── 5. Plot ───────────────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 5))

bars = ax.bar(categories, values, color=colors, edgecolor="white",
              linewidth=1.2, zorder=3)

# Value labels on bars
for bar, val in zip(bars, values):
    ax.text(bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.07,
            str(val),
            ha="center", va="bottom",
            fontsize=11, fontweight="bold", color="#333333")

# Grid and spines
ax.set_axisbelow(True)
ax.yaxis.grid(True, linestyle="--", linewidth=0.6, alpha=0.7, color="#CCCCCC")
ax.set_axisbelow(True)
for spine in ["top", "right", "left"]:
    ax.spines[spine].set_visible(False)
ax.spines["bottom"].set_color("#BBBBBB")
ax.tick_params(axis="x", length=0, pad=6)
ax.tick_params(axis="y", length=0)

# Axis labels and title
ax.set_xlabel("DID Design Classification", fontsize=11, labelpad=8,
              color="#444444")
ax.set_ylabel("Number of Papers", fontsize=11, labelpad=8, color="#444444")
ax.set_title("AER Papers by DID Design Classification\n"
             "(Waves 1 & 2 — 24 papers)",
             fontsize=13, fontweight="bold", color="#222222", pad=12)

ax.set_ylim(0, max(values) + 1.2)
ax.set_yticks(range(0, max(values) + 2))
plt.xticks(fontsize=10, color="#333333")
plt.yticks(fontsize=10, color="#333333")

# Legend: estimable designs vs edge-cases
legend_items = [
    mpatches.Patch(color=PALETTE["CLA"],  label="CLA  — did_multiplegt_dyn (controls req.)"),
    mpatches.Patch(color=PALETTE["SAD"],  label="SAD  — did_multiplegt_dyn"),
    mpatches.Patch(color=PALETTE["HAD"],  label="HAD  — did_had"),
    mpatches.Patch(color=PALETTE["SFSD"], label="SFSD — did_multiplegt_dyn (continuous)"),
    mpatches.Patch(color=PALETTE["SSFSD"],label="SSFSD — did_multiplegt_dyn"),
    mpatches.Patch(color=PALETTE["OTHER"],label="OTHER / unclassified"),
]
ax.legend(handles=legend_items, loc="upper right", fontsize=8,
          framealpha=0.85, edgecolor="#CCCCCC")

plt.tight_layout()
OUT = "/workspace/classification_barplot.png"
plt.savefig(OUT, dpi=160, bbox_inches="tight")
plt.close()
print(f"Saved: {OUT}")
print("Counts:", dict(zip(categories, values)))
