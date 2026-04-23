import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("C:/Users/Usuario/Documents/GitHub/twfe_survey/reports/tracker_master_consolidado.xlsx")
ws = wb["Master"]

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Find rows for the 4 papers
papers_to_update = {
    "Duggan": ("Sin data", "IMS Health data (ims0106all.dta) no incluida"),
    "Chaney": ("Sin data", "COMPUSTAT (compu_panel.dta), birth_date.dta no incluidos"),
    "Gentzkow": ("Sin data", "Datos electorales de uselectionatlas.org no incluidos"),
    "Brambilla": ("Sin data", "Datos firma argentina EIA/aduana no incluidos"),
}

updated = []
for row in range(6, 65):
    ref = ws.cell(row=row, column=3).value  # Column C = Referencia
    if ref:
        for key, (estado, nota) in papers_to_update.items():
            if key in str(ref):
                ws.cell(row=row, column=7).value = estado  # Column G = Estado
                ws.cell(row=row, column=7).fill = red_fill
                ws.cell(row=row, column=8).value = nota    # Column H = Notas
                updated.append(f"Row {row}: {ref} -> {estado}")
                break

for u in updated:
    print(u)

if len(updated) != 4:
    print(f"WARNING: Expected 4 updates, got {len(updated)}")
    # List all references to find mismatches
    for row in range(6, 65):
        ref = ws.cell(row=row, column=3).value
        if ref:
            print(f"  Row {row}: {ref}")

wb.save("C:/Users/Usuario/Documents/GitHub/twfe_survey/reports/tracker_master_consolidado.xlsx")
print("Excel saved successfully")
