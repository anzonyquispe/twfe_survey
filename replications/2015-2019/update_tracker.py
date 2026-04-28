"""
Update tracker_master_consolidado.xlsx with results from:
1. Munshi and Rosenzweig (2016) - No TWFE disponible
2. Dell (2015) - No TWFE disponible
3. Burgess et al. (2015) - Replicado, 0% neg weights
"""
import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(
    'C:/Users/Usuario/Documents/GitHub/twfe_survey/reports/tracker_master_consolidado.xlsx'
)
ws = wb.active

# Colors matching existing convention
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Column layout (from row 5):
# C1=#, C2=Ola, C3=Referencia, C4=Año, C5=Diseño, C6=Software, C7=Estado, C8=Notas

updates = {
    'dell': {
        'estado': 'No TWFE',
        'fill': yellow_fill,
        'notas': 'Sin TWFE disponible. Tables 1-5 son RD (regresion discontinuity). Tables 6-7 requieren datos confidenciales (drughom redactado).'
    },
    'burgess': {
        'estado': 'Replicado',
        'fill': blue_fill,
        'notas': '0% neg weights. Table 1 Col 1: areg exp_dens_share president i.year, absorb(distnum). beta=0.97 [0.36]. 319 pos, 0 neg weights. Treatment switches on/off.'
    },
    'munshi': {
        'estado': 'No TWFE',
        'fill': yellow_fill,
        'notas': 'Sin TWFE disponible. Datos cross-section (castes x villages, sin dimension temporal). Tables 6, 8a: OLS/areg/cgmwildboot.'
    }
}

for row in ws.iter_rows(min_row=6, max_col=8):
    ref = str(row[2].value).lower() if row[2].value else ''

    for key, update in updates.items():
        if key in ref:
            print(f"Updating row {row[0].row}: {row[2].value}")
            # Column 7 = Estado (index 6)
            row[6].value = update['estado']
            row[6].fill = update['fill']
            # Column 8 = Notas (index 7)
            row[7].value = update['notas']
            print(f"  Estado -> {update['estado']}")
            print(f"  Notas -> {update['notas'][:60]}...")
            break

wb.save('C:/Users/Usuario/Documents/GitHub/twfe_survey/reports/tracker_master_consolidado.xlsx')
print("\n=== Tracker saved successfully ===")
